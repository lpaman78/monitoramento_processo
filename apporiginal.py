
import threading
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import os
from glob import glob
import logging
import io
import warnings
from fpdf import FPDF
import win32com.client as win32
from tabulate import tabulate
from bs4 import BeautifulSoup
import time
import plotly.express as px
import streamlit.components.v1 as components
import plotly.graph_objects as go
import dash
from dash import dcc, html, Input, Output
import numpy as np
from scipy.stats import norm
import zipfile
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from datetime import datetime
import re
from io import BytesIO
import textwrap
from textwrap import wrap
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import matplotlib.animation as animation
from filelock import FileLock
from openpyxl import load_workbook
import calendar
import locale
from scipy.stats import shapiro
from PIL import Image


pd.set_option("styler.render.max_elements", 500000)  # ou um valor maior que o número de células
st.set_page_config(layout="wide")

# Função para capturar os logs
class StreamToLogger(io.StringIO):
    def __init__(self, logger, log_level=logging.INFO):
        super().__init__()
        self.logger = logger
        self.log_level = log_level

    def write(self, message):
        if message.strip():  # Evitar mensagens vazias
            self.logger.log(self.log_level, message.strip())

    def flush(self):
        pass

# Configurar logging para capturar os logs
log_capture_string = StreamToLogger(logging.getLogger(), logging.INFO)
logging.basicConfig(level=logging.INFO, handlers=[logging.StreamHandler(log_capture_string)])

# Suprimir avisos específicos
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')
warnings.filterwarnings("ignore", category=FutureWarning, module='pandas')
warnings.filterwarnings("ignore", category=FutureWarning, message="Setting an item of incompatible dtype is deprecated")

# Configurar logging
logging.basicConfig(level=logging.INFO)
CONSTANTES = {
    'VALOR_PADRAO_PRESSAO': None,
    'VALOR_PRESSAO_REBARBAS': None,
    'VALOR_PRESSAO_FALHA': None,
    'VALOR_PADRAO_TEMPO_INJ': None,
    'VALOR_TEMPO_INJ_REBARBAS': None,
    'VALOR_TEMPO_INJ_FALHA': None,
    'VALOR_PADRAO_ALMOFADA': None,
    'VALOR_ALMOFADA_REBARBAS': None,
    'VALOR_ALMOFADA_FALHA': None,
    'TOLERANCIA_ALMOFADA': None,
    'TOLERANCIA_PRESSAO': None,
    'TOLERANCIA_TEMPO_INJ': None,
    'LIMITE_ULTIMAS_LINHAS_FORA': None
}

contador_ciclos = 0
maquina_atual = 0  # Variável global para rastrear a máquina atual

def configurar_logging():
    """Configura o logging com formatação personalizada."""
    class CustomFormatter(logging.Formatter):
        def format(self, record):
            log_entry = super().format(record)
            if "Valor fora do intervalo" in log_entry:
                return f"\033[31m{log_entry}\033[0m"  # Adiciona cor vermelho escuro
            return log_entry

    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    file_handler = logging.FileHandler('monitoramento.log')
    file_handler.setLevel(logging.INFO)
    
    formatter = CustomFormatter('%(levelname)s - %(message)s')
    console_handler.setFormatter(formatter)
    file_handler.setFormatter(formatter)
    
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

configurar_logging()

def carregar_arquivo(caminho_pasta):
    """Carrega o arquivo Excel mais recente na pasta especificada e retorna o caminho do arquivo."""
    if not os.path.exists(caminho_pasta):
        logging.error(f"O caminho especificado não existe: {caminho_pasta}")
        return None
    
    arquivos_xlsx = glob(os.path.join(caminho_pasta, '*.xlsx'))
    
    if not arquivos_xlsx:
        logging.error("Nenhum arquivo encontrado na pasta.")
        return None
    
    arquivo_mais_recente = max(arquivos_xlsx, key=os.path.getmtime)
    logging.info(f"Arquivo mais recente carregado com sucesso: {os.path.basename(arquivo_mais_recente)}")
    
    return arquivo_mais_recente

def determinar_turno(data):
    hora = data.hour
    if 6 <= hora < 14:
        return "1° Turno"
    elif 14 <= hora < 22:
        return "2° Turno"
    else:
        return "3° Turno"
    
def formatar_nome_arquivo(nome):
    nome_formatado = nome.replace('_', ' ').replace('.xlsx', '')
    codigo_produto = nome_formatado.split(' ')[0]  # Extrair o código do produto antes do primeiro espaço
    return f"{codigo_produto}\n{nome_formatado[len(codigo_produto)+1:]}"

def processar_tabela(tabela):
    """
    Processa a tabela carregada:
    - Remove valores nulos e em branco nas colunas críticas
    - Renomeia coluna de injeções, se necessário
    - Converte datas do formato serial Excel para datetime
    - Trunca microsegundos para facilitar remoção de duplicatas
    - Elimina duplicatas mantendo a primeira ocorrência por data/hora
    - Ordena pela data mais antiga
    - Converte colunas numéricas
    - Renumera a coluna 'Número de injeção [StZx]' sequencialmente
    """

    colunas_necessarias = [
        'Data',
        'Número de injeção [StZx]',
        'Tempo de ciclo [Zus]',
        'Almofada [CPx]',
        'Troca de pressão [Phu]',
        'Tempo de injeção [ZSx]'
    ]

    # Renomear coluna de injeções, se necessário
    if 'Número total de injeções (SZX)' in tabela.columns:
        tabela = tabela.rename(columns={'Número total de injeções (SZX)': 'Número de injeção [StZx]'})

    # Substituir espaços em branco por NA
    tabela.replace(r'^\s*$', pd.NA, regex=True, inplace=True)

    # Remover linhas com todos os valores nulos nas colunas críticas
    colunas_existentes = [col for col in colunas_necessarias if col in tabela.columns]
    tabela = tabela.dropna(subset=colunas_existentes, how='all')

    # Converter datas do formato serial Excel
    if pd.api.types.is_numeric_dtype(tabela['Data']):
        tabela['Data'] = pd.to_timedelta(tabela['Data'], unit='d') + pd.Timestamp('1899-12-30')
    tabela['Data'] = pd.to_datetime(tabela['Data'], errors='coerce')
    tabela = tabela.dropna(subset=['Data'])

    # Truncar microsegundos para facilitar remoção de duplicatas
    tabela['Data'] = tabela['Data'].dt.floor('s')

    # Remover duplicatas mantendo apenas a primeira ocorrência
    tabela = tabela.drop_duplicates(subset=['Data'], keep='first')

    # Ordenar pela data
    tabela = tabela.sort_values(by='Data', ascending=True)

    # Converter colunas numéricas
    colunas_numericas = [
        'Número de injeção [StZx]',
        'Tempo de ciclo [Zus]',
        'Almofada [CPx]',
        'Troca de pressão [Phu]',
        'Tempo de injeção [ZSx]'
    ]
    for coluna in colunas_numericas:
        if coluna in tabela.columns:
            tabela[coluna] = pd.to_numeric(tabela[coluna], errors='coerce')

    # Renumerar a coluna de injeção
    tabela['Número de injeção [StZx]'] = range(1, len(tabela) + 1)

    return tabela











def subtrair_horarios(tabela):
    # Converter a coluna 'Data' para datetime
    tabela['Data'] = pd.to_datetime(tabela['Data'])
    
    # Inicializar lista para armazenar diferenças de tempo
    diferencas_tempo = []
    
    # Iterar sobre as linhas da tabela
    for i in range(1, len(tabela)):
        tempo_atual = tabela.iloc[i]['Data']
        tempo_anterior = tabela.iloc[i-1]['Data']
        diferenca_tempo = (tempo_anterior - tempo_atual).total_seconds()  # Subtrair a linha atual da linha anterior
        diferencas_tempo.append({
            'Linha Atual': i,
            'Linha Anterior': i-1,
            'Data Atual': tempo_atual,
            'Data Anterior': tempo_anterior,
            'Diferenca (segundos)': diferenca_tempo
        })
    
    # Remover o log das diferenças de tempo
    # for diferenca in diferencas_tempo:
    # logging.info(f"Linha Atual: {diferenca['Linha Atual']}, Linha Anterior: {diferenca['Linha Anterior']}, Data Atual: {diferenca['Data Atual']}, Data Anterior: {diferenca['Data Anterior']}, Diferenca (segundos): {diferenca['Diferenca (segundos)']}")
    
    return diferencas_tempo

def calcular_media_tempo_ciclo(tabela):
    """
    Função para calcular a média do tempo de ciclo na coluna 'Tempo de ciclo [Zus]',
    ignorando valores NaN e iguais a zero.
    Parâmetros:
        tabela (pd.DataFrame): DataFrame contendo a coluna 'Tempo de ciclo [Zus]'.
    Retorna:
        float: Média do tempo de ciclo.
    """
    coluna = 'Tempo de ciclo [Zus]'
    if coluna not in tabela.columns:
        raise ValueError(f"A coluna '{coluna}' não está presente na tabela.")

    # Filtra valores válidos (não nulos e diferentes de zero)
    valores_validos = tabela[coluna]
    valores_filtrados = valores_validos[(valores_validos.notna()) & (valores_validos != 0)]

    media_tempo_ciclo = valores_filtrados.mean()

    if not isinstance(media_tempo_ciclo, (float, int)):
        raise TypeError("A média do tempo de ciclo deve ser um valor numérico (float ou int).")

    return media_tempo_ciclo

def renomear_colunas_para_padrao(tabela):
    """Renomeia colunas específicas para padronização, se existirem."""
    colunas_para_renomear = {
        'Número total de injeções (SZX)': 'Número de injeção [StZx]'
    }

    colunas_existentes = set(tabela.columns)
    colunas_presentes_para_renomear = {
        original: novo for original, novo in colunas_para_renomear.items()
        if original in colunas_existentes
    }

    if colunas_presentes_para_renomear:
        tabela = tabela.rename(columns=colunas_presentes_para_renomear)
    else:
        logging.warning("Nenhuma das colunas a serem renomeadas foi encontrada.")

    return tabela

def remover_duplicatas(tabela):
    """Remove linhas duplicadas da coluna 'Data' e deixa apenas uma linha da informação."""
    return tabela.drop_duplicates(subset=['Data'])



 

    
def color_negative_red(val, valor_padrao, tolerancia, percentual=False):
    """Aplica cor vermelho escuro se o valor estiver fora do intervalo."""
    if isinstance(val, (int, float)):
        color = '\033[31m' if not valor_dentro_intervalo(val, valor_padrao, tolerancia, percentual) else ''
        return f'{color}{val}\033[0m' if color else val
    return val

def remover_colunas(tabela, colunas_para_remover):
    """Remove as colunas especificadas da tabela."""
    return tabela.drop(columns=colunas_para_remover, errors='ignore')

def filter_last_count(tabela, coluna):
    """Filtra as contagens na coluna especificada para manter apenas a última contagem completa."""
    ultima_contagem = tabela[coluna].max()
    tabela_filtrada = tabela[tabela[coluna] <= ultima_contagem]
    return tabela_filtrada

def formatar_coluna_troca_pressao(tabela):
    """Formata a coluna 'Troca de pressão [Phu]' com duas casas decimais."""
    if 'Troca de pressão [Phu]' in tabela.columns:
        tabela['Troca de pressão [Phu]'] = tabela['Troca de pressão [Phu]'].map('{:.2f}'.format)
    return tabela

def formatar_coluna_almofada(tabela):
    """Formata a coluna 'Almofada [CPx]' com duas casas decimais."""
    if 'Almofada [CPx]' in tabela.columns:
        tabela['Almofada [CPx]'] = tabela['Almofada [CPx]'].map('{:.2f}'.format)
    return tabela

def formatar_coluna_tempo_injecao(tabela):
    """Formata a coluna 'Tempo de injeção [ZSx]' com duas casas decimais."""
    if 'Tempo de injeção [ZSx]' in tabela.columns:
        tabela['Tempo de injeção [ZSx]'] = tabela['Tempo de injeção [ZSx]'].map('{:.2f}'.format)
    return tabela

def ajustar_segundos_ciclo(tabela):
    """Ajusta os segundos do ciclo igual ao que está na tabela."""
    colunas_para_ajustar = ['Tempo de ciclo [Zus]', 'Tempo de injeção [ZSx]']
    
    for coluna in colunas_para_ajustar:
        if coluna in tabela.columns:
            tabela[coluna] = tabela[coluna].round(2)
        else:
            logging.warning(f"Coluna '{coluna}' não encontrada na tabela.")
    
    return tabela

def calcular_medias(tabela):
    """
    Função para calcular as médias dos parâmetros de injeção, almofada e pressão,
    ignorando valores NaN e iguais a zero.
    Parâmetros:
        tabela (pd.DataFrame): DataFrame contendo as colunas dos parâmetros.
    Retorna:
        dict: Dicionário com as médias calculadas.
    """
    medias = {
        'VALOR_PADRAO_TEMPO_INJ': tabela['Tempo de injeção [ZSx]'][(tabela['Tempo de injeção [ZSx]'].notna()) & (tabela['Tempo de injeção [ZSx]'] != 0)].mean(),
        'VALOR_PADRAO_ALMOFADA': tabela['Almofada [CPx]'][(tabela['Almofada [CPx]'].notna()) & (tabela['Almofada [CPx]'] != 0)].mean(),
        'VALOR_PADRAO_PRESSAO': tabela['Troca de pressão [Phu]'][(tabela['Troca de pressão [Phu]'].notna()) & (tabela['Troca de pressão [Phu]'] != 0)].mean()
    }
    return medias

def converter_horas_centesimais_para_horas_normais(horas_centesimais):
    horas = int(horas_centesimais)
    minutos = int((horas_centesimais - horas) * 60)
    segundos = int((((horas_centesimais - horas) * 60) - minutos) * 60)
    return f"{horas:02d}:{minutos:02d}:{segundos:02d}"

def verificar_paradas_maquina(tabela, media_tempo_ciclo):
    """
    Verifica paradas de máquina com base na diferença de tempo entre registros ('Intervalo')
    e na média do tempo de ciclo multiplicada por cinco.

    Parâmetros:
        tabela (pd.DataFrame): DataFrame contendo a coluna 'Data'.
        media_tempo_ciclo (float): Média do tempo de ciclo.

    Retorna:
        dict: Dicionário contendo informações sobre as paradas de máquina por turno.
    """
    if not isinstance(media_tempo_ciclo, (float, int)):
        raise TypeError("A média do tempo de ciclo deve ser um valor numérico (float ou int).")

    # Garantir que a coluna 'Data' esteja em datetime e ordenada
    tabela = tabela.copy()
    tabela['Data'] = pd.to_datetime(tabela['Data'], errors='coerce')
    tabela = tabela.dropna(subset=['Data'])
    tabela = tabela.sort_values(by='Data')

    # Calcular intervalo entre registros
    tabela['Intervalo'] = tabela['Data'].diff().dt.total_seconds()

    paradas_por_turno = {turno: [] for turno in ["1° Turno", "2° Turno", "3° Turno"]}

    for i in range(1, len(tabela)):
        linha = tabela.iloc[i]
        intervalo = linha['Intervalo']
        data_atual = linha['Data']

        if pd.notna(intervalo) and intervalo > media_tempo_ciclo * 5:
            duracao_minutos = (intervalo - media_tempo_ciclo * 2) / 60
            duracao_horas = duracao_minutos / 60
            duracao_formatada = converter_horas_centesimais_para_horas_normais(duracao_horas)
            turno = determinar_turno(data_atual)
            paradas_por_turno[turno].append({
                'Data': data_atual,
                'Duração (minutos)': round(duracao_minutos, 2),
                'Duração formatada': duracao_formatada,
                'Turno': turno
            })

    return paradas_por_turno



def atualizar_parametros(caminho_parametros, maquina, produto, medias):
    """Atualiza a planilha de parâmetros com as médias calculadas, se a máquina e o produto não existirem."""
    if not os.path.exists(caminho_parametros):
        logging.error(f"O arquivo de parâmetros não foi encontrado: {caminho_parametros}")
        return None

    parametros_df = pd.read_excel(caminho_parametros, engine='openpyxl')
    parametros_df['MAQUINA'] = parametros_df['MAQUINA'].str.strip().astype(str)
    parametros_df['PRODUTO'] = parametros_df['PRODUTO'].astype(str).str.zfill(9)  # Garantir que todos os códigos de produto tenham 9 dígitos

    maquina = maquina.strip()
    produto = str(produto).strip().zfill(9)  # Garantir que o código do produto tenha 9 dígitos, preenchendo com zeros à esquerda se necessário

    # Verificar se a máquina e o produto já existem na planilha
    if ((parametros_df['MAQUINA'] == maquina) & (parametros_df['PRODUTO'] == produto)).any():
        # Atualizar os valores existentes
        parametros_df.loc[(parametros_df['MAQUINA'] == maquina) & (parametros_df['PRODUTO'] == produto), [
            'VALOR_PADRAO_PRESSAO', 'VALOR_PADRAO_TEMPO_INJ', 'VALOR_PADRAO_ALMOFADA',
            'VALOR_PRESSAO_REBARBAS', 'VALOR_PRESSAO_FALHA', 'VALOR_TEMPO_INJ_REBARBAS',
            'VALOR_TEMPO_INJ_FALHA', 'VALOR_ALMOFADA_REBARBAS', 'VALOR_ALMOFADA_FALHA',
            'TOLERANCIA_ALMOFADA', 'TOLERANCIA_PRESSAO', 'TOLERANCIA_TEMPO_INJ'
        ]] = [
            round(medias['VALOR_PADRAO_PRESSAO'], 2),
            round(medias['VALOR_PADRAO_TEMPO_INJ'], 2),
            round(medias['VALOR_PADRAO_ALMOFADA'], 2),
            round(medias['VALOR_PADRAO_PRESSAO'] + medias['VALOR_PADRAO_PRESSAO'] * 0.30, 2),
            round(medias['VALOR_PADRAO_PRESSAO'] - medias['VALOR_PADRAO_PRESSAO'] * 0.30, 2),
            round(medias['VALOR_PADRAO_TEMPO_INJ'] + medias['VALOR_PADRAO_TEMPO_INJ'] * 0.15, 2),
            round(medias['VALOR_PADRAO_TEMPO_INJ'] - medias['VALOR_PADRAO_TEMPO_INJ'] * 0.15, 2),
            round(medias['VALOR_PADRAO_ALMOFADA'] + medias['VALOR_PADRAO_ALMOFADA'] * 0.30, 2),
            round(medias['VALOR_PADRAO_ALMOFADA'] - medias['VALOR_PADRAO_ALMOFADA'] * 0.30, 2),
            round(2.00, 2),  # Valor fixo de 2,00 na planilha
            round(0.15, 2),  # Valor fixo de 0,10 na planilha
            round(0.15, 2)   # Valor fixo de 0,10 na planilha
        ]
        logging.info(f"Parâmetros atualizados com sucesso para a máquina {maquina} e produto {produto}.")
    else:
        # Adicionar nova linha com as médias calculadas
        nova_linha = pd.DataFrame([{
            'MAQUINA': maquina,
            'PRODUTO': produto,
            'VALOR_PADRAO_PRESSAO': round(medias['VALOR_PADRAO_PRESSAO'], 2),
            'VALOR_PADRAO_TEMPO_INJ': round(medias['VALOR_PADRAO_TEMPO_INJ'], 2),
            'VALOR_PADRAO_ALMOFADA': round(medias['VALOR_PADRAO_ALMOFADA'], 2),
            'VALOR_PRESSAO_REBARBAS': round(medias['VALOR_PADRAO_PRESSAO'] + medias['VALOR_PADRAO_PRESSAO'] * 0.30, 2),
            'VALOR_PRESSAO_FALHA': round(medias['VALOR_PADRAO_PRESSAO'] - medias['VALOR_PADRAO_PRESSAO'] * 0.30, 2),
            'VALOR_TEMPO_INJ_REBARBAS': round(medias['VALOR_PADRAO_TEMPO_INJ'] + medias['VALOR_PADRAO_TEMPO_INJ'] * 0.15, 2),
            'VALOR_TEMPO_INJ_FALHA': round(medias['VALOR_PADRAO_TEMPO_INJ'] - medias['VALOR_PADRAO_TEMPO_INJ'] * 0.15, 2),
            'VALOR_ALMOFADA_REBARBAS': round(medias['VALOR_PADRAO_ALMOFADA'] + medias['VALOR_PADRAO_ALMOFADA'] * 0.30, 2),
            'VALOR_ALMOFADA_FALHA': round(medias['VALOR_PADRAO_ALMOFADA'] - medias['VALOR_PADRAO_ALMOFADA'] * 0.30, 2),
            'TOLERANCIA_ALMOFADA': round(2.00, 2),  # Valor fixo de 2,00 na planilha
            'TOLERANCIA_PRESSAO': round(0.15, 2),   # Valor fixo de 0,10 na planilha
            'TOLERANCIA_TEMPO_INJ': round(0.15, 2), # Valor fixo de 0,10 na planilha
            'LIMITE_ULTIMAS_LINHAS_FORA': 20  # Valor Padrão
        }])
        parametros_df = pd.concat([parametros_df, nova_linha], ignore_index=True)
        logging.info(f"Parâmetros adicionados com sucesso para a máquina {maquina} e produto {produto}.")
    
    parametros_df.to_excel(caminho_parametros, index=False, engine='openpyxl')

# Exemplo de uso
medias = {
    'VALOR_PADRAO_PRESSAO': 77.06,
    'VALOR_PADRAO_TEMPO_INJ': 1.50,
    'VALOR_PADRAO_ALMOFADA': 10.20
}
atualizar_parametros('caminho_parametros.xlsx', 'Maquina1', '018178001', medias)

def carregar_parametros(caminho_parametros, codigo_produto, numero_maquina):
    """Carrega os parâmetros do processo a partir de um arquivo padrão."""
    if not os.path.exists(caminho_parametros):
        logging.error(f"O arquivo de parâmetros não foi encontrado: {caminho_parametros}")
        return None

    parametros_df = pd.read_excel(caminho_parametros, engine='openpyxl')
    parametros_df['MAQUINA'] = parametros_df['MAQUINA'].str.strip().astype(str)
    parametros_df['PRODUTO'] = parametros_df['PRODUTO'].astype(str).str.zfill(9)  # Garantir que todos os códigos de produto tenham 9 dígitos

    numero_maquina = numero_maquina.strip()
    codigo_produto = str(codigo_produto).strip().zfill(9)  # Garantir que o código do produto tenha 9 dígitos, preenchendo com zeros à esquerda se necessário

    logging.info(f"DataFrame de parâmetros carregado: {parametros_df.head()}")  # Adicionar log para verificar o DataFrame carregado
    logging.info(f"Carregando parâmetros para o código do produto: {codigo_produto} e número da máquina: {numero_maquina}")

    parametros_df = parametros_df[(parametros_df['MAQUINA'] == numero_maquina) & (parametros_df['PRODUTO'] == codigo_produto)]

    if parametros_df.empty:
        logging.error(f"Parâmetros não encontrados para o código do produto {codigo_produto} e número da máquina {numero_maquina}.")
        return None

    parametros = parametros_df.iloc[0].to_dict()
    return parametros

# Exemplo de uso
parametros = carregar_parametros('caminho_parametros.xlsx', '018178001', 'Maquina1')
print(parametros)

def valor_dentro_intervalo(valor, valor_padrao, tolerancia, percentual=False):
    """
    Verifica se o valor está dentro do intervalo permitido.
    Args:
        valor (float): Valor atual do parâmetro.
        valor_padrao (float): Valor de referência.
        tolerancia (float): Tolerância permitida.
        percentual (bool): Se True, aplica tolerância percentual.
    Returns:
        bool: True se estiver dentro do intervalo, False caso contrário.
    """
    # Verificações básicas
    if valor is None or valor_padrao is None or tolerancia is None:
        return False

    if not isinstance(valor, (int, float)) or not isinstance(valor_padrao, (int, float)) or not isinstance(tolerancia, (int, float)):
        return False

    # Cálculo com tolerância percentual
    if percentual:
        limite_inferior = valor_padrao * (1 - tolerancia)
        limite_superior = valor_padrao * (1 + tolerancia)
    else:
        limite_inferior = valor_padrao - tolerancia
        limite_superior = valor_padrao + tolerancia

    return limite_inferior <= valor <= limite_superior

def enviar_email_outlook(destinatario, assunto, corpo):
    """Envia email via Outlook."""
    try:
        outlook = win32.Dispatch('outlook.application')
        email = outlook.CreateItem(0)
        email.To = destinatario
        email.Subject = assunto
        email.HTMLBody = corpo
        email.Send()
        logging.info("E-mail enviado com sucesso!")
    except Exception as e:
        logging.error(f"Erro ao enviar email: {e}", exc_info=True)

# Função para formatar os nomes dos arquivos
def formatar_nome_maquina(maquina):
    partes = maquina.split('_')
    if len(partes) > 2:
        return f"Máquina {partes[1]} {partes[2]}"
    return maquina

def verificar_ciclos_por_turno(tabela, parametros, media_tempo_ciclo):
    turnos = {
        "1° Turno": ("06:00:00", "14:00:00"),
        "2° Turno": ("14:00:00", "22:00:00"),
        "3° Turno": ("22:00:00", "06:00:00")
    }

    tabela['Data'] = pd.to_datetime(tabela['Data'])
    tabela['Horario'] = tabela['Data'].dt.strftime('%H:%M:%S')

    def determinar_turno(horario):
        for turno, (inicio, fim) in turnos.items():
            if inicio <= horario < fim:
                return turno
        return "3° Turno"

    tabela['Turno'] = tabela['Horario'].apply(determinar_turno)
    tabela['DataAnterior'] = tabela['Data'].shift(1)
    tabela['DiferencaTempo'] = (tabela['Data'] - tabela['DataAnterior']).dt.total_seconds()

    ciclos_fora_total = 0
    ciclos_fora_por_turno = {turno: 0 for turno in turnos}
    total_ciclos_monitorados_por_turno = {turno: 0 for turno in turnos}
    parametros_fora = {turno: [] for turno in turnos}
    observacoes_por_turno = {turno: "Processo não está sendo Monitorado ou as Tolerâncias estão muito abertas. " for turno in turnos}
    ciclos_fora_consecutivos = {turno: 0 for turno in turnos}

    parametros_monitoramento = {
        'Almofada [CPx]': (parametros['VALOR_PADRAO_ALMOFADA'], parametros['TOLERANCIA_ALMOFADA'], False),
        'Troca de pressão [Phu]': (parametros['VALOR_PADRAO_PRESSAO'], parametros['TOLERANCIA_PRESSAO'], True),
        'Tempo de injeção [ZSx]': (parametros['VALOR_PADRAO_TEMPO_INJ'], parametros['TOLERANCIA_TEMPO_INJ'], False),
    }

    for index, row in tabela.iterrows():
        turno = row['Turno']
        total_ciclos_monitorados_por_turno[turno] += 1
        ciclo_fora = False

        for parametro, (valor_padrao, tolerancia, percentual) in parametros_monitoramento.items():
            valor = row[parametro]
            if pd.notna(valor):  # ✅ Ignora apenas NaN
                if not valor_dentro_intervalo(valor, valor_padrao, tolerancia, percentual):
                    ciclos_fora_total += 1
                    ciclos_fora_por_turno[turno] += 1
                    parametros_fora[turno].append(parametro)
                    ciclo_fora = True

        if ciclo_fora:
            ciclos_fora_consecutivos[turno] += 1
            diferenca_tempo = row['DiferencaTempo']
            if pd.notna(diferenca_tempo) and diferenca_tempo > media_tempo_ciclo:
                observacoes_por_turno[turno] = "Processo está sendo Monitorado adequadamente"
        else:
            ciclos_fora_consecutivos[turno] = 0

        if ciclos_fora_consecutivos[turno] >= 20:
            observacoes_por_turno[turno] = "Processo não está sendo Monitorado ou as Tolerâncias estão muito abertas. "

    for turno in turnos:
        if ciclos_fora_por_turno[turno] == 0:
            observacoes_por_turno[turno] = "Processo está dentro do especificado"

    percentual_ciclos_fora_por_turno = {
        turno: (ciclos_fora_por_turno[turno] / total_ciclos_monitorados_por_turno[turno]) * 100
        if total_ciclos_monitorados_por_turno[turno] > 0 else 0
        for turno in turnos
    }

    return (
        total_ciclos_monitorados_por_turno,
        ciclos_fora_total,
        percentual_ciclos_fora_por_turno,
        ciclos_fora_por_turno,
        parametros_fora,
        observacoes_por_turno
    )

def validar_valores_para_calculo(valores):
    """Verifica se todos os valores são válidos para cálculo estatístico (numéricos, finitos e não nulos). Zeros são permitidos."""
    return all(
        isinstance(v, (int, float)) and not pd.isna(v) and np.isfinite(v)
        for v in valores
    )

def calcular_cp_cpk(tabela, coluna, usl, lsl):
    """Calcula os índices Cp e Cpk para uma coluna de dados, ignorando NaN mas mantendo zeros."""
    if coluna not in tabela.columns or tabela[coluna].isnull().all():
        logging.error(f"Coluna {coluna} não encontrada ou sem dados.")
        return None, None

    tabela[coluna] = pd.to_numeric(tabela[coluna], errors='coerce')
    tabela = tabela.dropna(subset=[coluna])  # Mantém zeros

    if tabela[coluna].count() < 2:
        logging.warning(f"Coluna {coluna} não possui dados suficientes para cálculo.")
        return None, None

    media = tabela[coluna].mean()
    desvio_padrao = tabela[coluna].std(ddof=1)

    if not validar_valores_para_calculo([media, desvio_padrao, usl, lsl]):
        logging.error(f"Valores inválidos para cálculo de Cp/Cpk: {coluna}")
        return None, None

    if desvio_padrao == 0:
        logging.error(f"Desvio padrão é zero para a coluna {coluna}.")
        return None, None

    cp = (usl - lsl) / (6 * desvio_padrao)
    cpk = min((usl - media) / (3 * desvio_padrao), (media - lsl) / (3 * desvio_padrao))

    return cp, cpk

def calcular_pp_ppk(tabela, coluna, usl, lsl):
    """Calcula os índices Pp e Ppk para uma coluna de dados, ignorando NaN mas mantendo zeros."""
    if coluna not in tabela.columns or tabela[coluna].isnull().all():
        logging.error(f"Coluna {coluna} não encontrada ou sem dados.")
        return None, None

    tabela[coluna] = pd.to_numeric(tabela[coluna], errors='coerce')
    tabela = tabela.dropna(subset=[coluna])  # Mantém zeros

    if tabela[coluna].count() < 2:
        logging.warning(f"Coluna {coluna} não possui dados suficientes para cálculo.")
        return None, None

    media = tabela[coluna].mean()
    desvio_padrao = tabela[coluna].std(ddof=0)  # Desvio padrão populacional

    if not validar_valores_para_calculo([media, desvio_padrao, usl, lsl]):
        logging.error(f"Valores inválidos para cálculo de Pp/Ppk: {coluna}")
        return None, None

    if desvio_padrao == 0:
        logging.error(f"Desvio padrão é zero para a coluna {coluna}.")
        return None, None

    pp = (usl - lsl) / (6 * desvio_padrao)
    ppk = min((usl - media) / (3 * desvio_padrao), (media - lsl) / (3 * desvio_padrao))

    return pp, ppk

def calcular_cm_cmk(tabela, coluna, usl, lsl):
    """Calcula os índices Cm e Cmk para uma coluna de dados, ignorando NaN mas mantendo zeros."""
    if coluna not in tabela.columns or tabela[coluna].isnull().all():
        logging.error(f"Coluna {coluna} não encontrada ou sem dados.")
        return None, None

    tabela[coluna] = pd.to_numeric(tabela[coluna], errors='coerce')
    tabela = tabela.dropna(subset=[coluna])  # Mantém zeros

    if tabela[coluna].count() < 2:
        logging.warning(f"Coluna {coluna} não possui dados suficientes para cálculo.")
        return None, None

    media = tabela[coluna].mean()
    desvio_padrao = tabela[coluna].std(ddof=0)  # Desvio padrão populacional

    if not validar_valores_para_calculo([media, desvio_padrao, usl, lsl]):
        logging.error(f"Valores inválidos para cálculo de Cm/Cmk: {coluna}")
        return None, None

    if desvio_padrao == 0:
        logging.error(f"Desvio padrão é zero para a coluna {coluna}.")
        return None, None

    cm = (usl - lsl) / (6 * desvio_padrao)
    cmk = min((usl - media) / (3 * desvio_padrao), (media - lsl) / (3 * desvio_padrao))

    return cm, cmk

def verificar_normalidade(dados, alpha=0.05):
    """
    Verifica se os dados seguem uma distribuição normal usando o teste de Shapiro-Wilk.

    Parâmetros:
    - dados (iterável): Série ou lista de valores numéricos.
    - alpha (float): Nível de significância para o teste. Padrão é 0.05.

    Retorna:
    - (bool, float): 
        - True se os dados forem normalmente distribuídos (p-value > alpha), False caso contrário.
        - p-value do teste, ou None se os dados forem insuficientes.
    """
    dados = pd.to_numeric(dados, errors='coerce').dropna()
    if len(dados) < 3:
        return False, None  # Dados insuficientes para o teste
    stat, p_value = shapiro(dados)
    return p_value > alpha, p_value


def calcular_indices_com_normalidade(tabela, coluna, usl, lsl):
    """
    Calcula Cp/Cpk se normal, Pp/Ppk se não normal, e Cm/Cmk sempre.
    Retorna todos os índices e o tipo de distribuição.
    """
    # Validação dos limites
    try:
        usl = float(usl)
        lsl = float(lsl)
        if usl <= lsl:
            logging.warning(f"USL ({usl}) deve ser maior que LSL ({lsl}).")
            return None
    except (ValueError, TypeError):
        logging.warning(f"USL e LSL devem ser valores numéricos válidos. Recebido: USL={usl}, LSL={lsl}")
        return None

    dados = pd.to_numeric(tabela[coluna], errors='coerce').dropna()
    if len(dados) < 2:
        logging.warning(f"Coluna {coluna} não possui dados suficientes para cálculo.")
        return None

    if dados.std() == 0:
        logging.warning(f"Dados constantes na coluna {coluna}.")
        return None

    is_normal, p_value = verificar_normalidade(dados)

    if is_normal:
        cp, cpk = calcular_cp_cpk(tabela, coluna, usl, lsl)
        pp, ppk = None, None
        tipo = "Normal"
    else:
        pp, ppk = calcular_pp_ppk(tabela, coluna, usl, lsl)
        cp, cpk = None, None
        tipo = "Não Normal"

    cm, cmk = calcular_cm_cmk(tabela, coluna, usl, lsl)

    return {
        "Cp": cp,
        "Cpk": cpk,
        "Pp": pp,
        "Ppk": ppk,
        "Cm": cm,
        "Cmk": cmk,
        "Distribuição": tipo,
        "p-value": p_value
    }









def carregar_dados_todas_maquinas(maquinas, caminho_parametros):
    import os
    import logging
    import pandas as pd
    from glob import glob

    colunas_para_remover = [
        'Tempo de dosagem [ZDx]',
        'Número de peças boas injetadas [FTZX]',
        'Troca de posição [C3U]'
    ]
    historico_dados = []

    for maquina in maquinas:
        caminho_pasta = os.path.join(
            r"L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo",
            maquina
        )
        arquivos_xlsx = glob(os.path.join(caminho_pasta, '*.xlsx'))

        for arquivo in arquivos_xlsx:
            nome_arquivo = os.path.basename(arquivo)
            if nome_arquivo.startswith('~$') or os.path.getsize(arquivo) == 0:
                continue

            try:
                tabela = pd.read_excel(arquivo, engine='openpyxl')
                tabela = processar_tabela(tabela)
                if tabela is None:
                    continue

                tabela = ajustar_segundos_ciclo(tabela)
                tabela = remover_duplicatas(tabela)
                tabela = remover_colunas(tabela, colunas_para_remover)
                tabela = filter_last_count(tabela, 'Data')

                for coluna in ['Almofada [CPx]', 'Troca de pressão [Phu]', 'Tempo de injeção [ZSx]']:
                    if coluna in tabela.columns:
                        tabela[coluna] = pd.to_numeric(tabela[coluna], errors='coerce')
                        tabela = tabela[tabela[coluna].notna() & (tabela[coluna] != 0)]

                try:
                    tabela['Data'] = pd.to_datetime(tabela['Data'], errors='coerce').dt.date
                    data_arquivo = tabela['Data'].dropna().iloc[-1] if not tabela['Data'].isna().all() else None
                except Exception as e:
                    logging.warning(f"Erro ao extrair data do arquivo {nome_arquivo}: {e}")
                    data_arquivo = None

                codigo_produto = nome_arquivo.split('_')[0]
                parametros = carregar_parametros(caminho_parametros, codigo_produto, maquina)
                if parametros is None:
                    continue

                limites = {
                    'Almofada [CPx]': {'usl': parametros['VALOR_ALMOFADA_REBARBAS'], 'lsl': parametros['VALOR_ALMOFADA_FALHA']},
                    'Troca de pressão [Phu]': {'usl': parametros['VALOR_PRESSAO_REBARBAS'], 'lsl': parametros['VALOR_PRESSAO_FALHA']},
                    'Tempo de injeção [ZSx]': {'usl': parametros['VALOR_TEMPO_INJ_REBARBAS'], 'lsl': parametros['VALOR_TEMPO_INJ_FALHA']}
                }

                cp_cpk = exibir_cp_cpk_pp_ppk(tabela, limites)
                cm_cmk = exibir_cm_cmk(tabela, limites)
                media_ciclo = calcular_media_tempo_ciclo(tabela)
                ciclos_por_turno, ciclos_fora, *_ = verificar_ciclos_por_turno(tabela, parametros, media_ciclo)
                paradas_por_turno = verificar_paradas_maquina(tabela, media_ciclo)
                total_ciclos = sum(ciclos_por_turno.values())
                total_paradas = sum(len(paradas_por_turno[t]) for t in paradas_por_turno)

                def extrair_valor(linha):
                    return linha.split(": ")[1] if "N/A" not in linha else None

                estatisticas_cp_cpk = {
                    f'{indice}_{param}': extrair_valor(cp_cpk[i][j])
                    for i, param in enumerate(['Almofada', 'Pressao', 'Tempo_Inj'])
                    for j, indice in enumerate(['Cp', 'Cpk', 'Pp', 'Ppk'])
                }

                estatisticas_cm_cmk = {
                    f'{indice}_{param}': extrair_valor(cm_cmk[i][j])
                    for i, param in enumerate(['Almofada', 'Pressao', 'Tempo_Inj'])
                    for j, indice in enumerate(['Cm', 'Cmk'])
                }

                estatisticas_combinadas = {**estatisticas_cp_cpk, **estatisticas_cm_cmk}

                if all(v in [None, ''] for v in estatisticas_combinadas.values()):
                    logging.info(f"Linha ignorada por ausência de valores estatísticos: {nome_arquivo}")
                    continue

                historico_dados.append({
                    'Arquivo': nome_arquivo,
                    'Maquina': maquina,
                    'Data': data_arquivo,
                    'Ordem de Produção': nome_arquivo.split('_')[1].replace('.xlsx', '') if '_' in nome_arquivo else None,
                    'Part Number': codigo_produto.lstrip('0'),
                    **estatisticas_combinadas,
                    'Total_Ciclos_Monitorados': total_ciclos,
                    'Total_Ciclos_Fora': ciclos_fora,
                    'Total_Paradas_Maquina': total_paradas
                })

            except Exception as e:
                logging.error(f"Erro ao processar {nome_arquivo} da máquina {maquina}: {e}", exc_info=True)

    return historico_dados

def atualizar_resultados_estatisticos(caminho_arquivo, nova_linha):
    import pandas as pd
    import os
    from filelock import FileLock
    import logging
    from datetime import datetime

    colunas_estatisticas = [
        "Cp_Tempo_Inj", "Cpk_Tempo_Inj", "Pp_Tempo_Inj", "Ppk_Tempo_Inj", "Cm_Tempo_Inj", "Cmk_Tempo_Inj",
        "Cp_Almofada", "Cpk_Almofada", "Pp_Almofada", "Ppk_Almofada", "Cm_Almofada", "Cmk_Almofada",
        "Cp_Pressao", "Cpk_Pressao", "Pp_Pressao", "Ppk_Pressao", "Cm_Pressao", "Cmk_Pressao"
    ]

    if all(nova_linha.get(col) in [None, ""] for col in colunas_estatisticas):
        logging.info("Todos os valores estatísticos estão ausentes ou em branco. Nenhuma atualização realizada.")
        return

    dir_path = os.path.dirname(caminho_arquivo)
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)

    lock_path = caminho_arquivo + ".lock"
    with FileLock(lock_path, timeout=10):
        if os.path.exists(caminho_arquivo):
            try:
                df_existente = pd.read_excel(caminho_arquivo, engine="openpyxl", dtype=str)
            except Exception:
                df_existente = pd.DataFrame(columns=nova_linha.keys())
        else:
            df_existente = pd.DataFrame(columns=nova_linha.keys())

        # Normalização
        for campo in ["Ordem de Produção", "Maquina", "Part Number"]:
            if campo in df_existente.columns:
                df_existente[campo] = df_existente[campo].astype(str).str.strip()
            nova_linha[campo] = str(nova_linha.get(campo, "")).strip()

        # Garantir que 'Data' esteja presente e formatada
        if "Data" not in nova_linha or not nova_linha["Data"]:
            nova_linha["Data"] = datetime.today().date()
        else:
            nova_linha["Data"] = pd.to_datetime(nova_linha["Data"], errors="coerce").date()

        # Ordenar por data decrescente
        df_existente["Data"] = pd.to_datetime(df_existente["Data"], errors="coerce").dt.date
        df_existente = df_existente.sort_values(by="Data", ascending=False)

        # Remover duplicatas mantendo a mais recente
        df_existente = df_existente.drop_duplicates(
            subset=["Ordem de Produção", "Maquina", "Part Number"], keep="first"
        )

        # Atualizar ou adicionar nova linha
        mask = (
            (df_existente["Ordem de Produção"] == nova_linha["Ordem de Produção"]) &
            (df_existente["Maquina"] == nova_linha["Maquina"]) &
            (df_existente["Part Number"] == nova_linha["Part Number"])
        )

        if mask.any():
            for coluna, valor in nova_linha.items():
                if valor not in [None, ""]:
                    df_existente.loc[mask, coluna] = valor
            logging.info("Linha existente atualizada.")
        else:
            df_existente = pd.concat([df_existente, pd.DataFrame([nova_linha])], ignore_index=True)
            logging.info("Nova linha adicionada ao histórico.")

        df_existente.drop(columns=["Arquivo"], errors="ignore", inplace=True)

        temp_path = caminho_arquivo.replace(".xlsx", "_temp.xlsx")
        df_existente.to_excel(temp_path, index=False, engine="openpyxl")
        os.replace(temp_path, caminho_arquivo)
        logging.info(f"Histórico atualizado com sucesso: {caminho_arquivo}")

def obter_nome_maquina_com_status(maquina, caminho_parametros):
    nome_formatado = formatar_nome_maquina(maquina)
    caminho_pasta = os.path.join(
        "L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo",
        maquina
    )

    # 1. Carregamento do arquivo
    caminho_arquivo = carregar_arquivo(caminho_pasta)
    if not caminho_arquivo:
        return f"📄 {nome_formatado} Dados não encontrados"

    try:
        tabela = pd.read_excel(caminho_arquivo, engine='openpyxl')
    except Exception as e:
        return f"📄 {nome_formatado} Erro ao ler os Dados: {e}"

    # 2. Processamento da tabela
    try:
        tabela = processar_tabela(tabela)
        tabela = ajustar_segundos_ciclo(tabela)
        tabela = remover_duplicatas(tabela)
        tabela = filter_last_count(tabela, 'Data')
    except Exception as e:
        return f"📉 {nome_formatado} Erro ao processar Tabela: {e}"

    # 3. Extração do código do produto
    codigo_produto = os.path.basename(caminho_arquivo).split('_')[0].strip().zfill(9)

    # 4. Carregamento dos parâmetros
    parametros = carregar_parametros(caminho_parametros, codigo_produto, maquina)
    if parametros is None:
        return f"⚙️ {nome_formatado} Parâmetros não encontrados"

    # 5. Definição dos limites
    limites = {
        'Almofada [CPx]': {
            'usl': parametros['VALOR_ALMOFADA_REBARBAS'],
            'lsl': parametros['VALOR_ALMOFADA_FALHA']
        },
        'Troca de pressão [Phu]': {
            'usl': parametros['VALOR_PRESSAO_REBARBAS'],
            'lsl': parametros['VALOR_PRESSAO_FALHA']
        },
        'Tempo de injeção [ZSx]': {
            'usl': parametros['VALOR_TEMPO_INJ_REBARBAS'],
            'lsl': parametros['VALOR_TEMPO_INJ_FALHA']
        }
    }

    # 6. Cálculo dos índices
    try:
        cp_cpk_pp_ppk_linhas = exibir_cp_cpk_pp_ppk(tabela, limites)
        cm_cmk_linhas = exibir_cm_cmk(tabela, limites)
    except Exception as e:
        return f"📉 {nome_formatado} Erro ao calcular índices: {e}"

    # 7. Análise dos resultados
    analise_cp_cpk = analisar_resultados_cp_cpk_pp_ppk(cp_cpk_pp_ppk_linhas)
    analise_cm_cmk = analisar_resultados_cm_cmk(cm_cmk_linhas)
    todas_analises = analise_cp_cpk + analise_cm_cmk

    # 8. Determinação do símbolo
    simbolo = "✅" if all("✅" in linha for linha in todas_analises) else (
              "❌" if any("❌" in linha for linha in todas_analises) else "⚠️")

    return f"{simbolo} {nome_formatado}"







def analisar_resultados_cp_cpk_pp_ppk(cp_cpk_pp_ppk_linhas):
    """Analisa os resultados de Cp, Cpk, Pp e Ppk e retorna a análise com cores no Streamlit."""
    parametros = ["Almofada [CPx]", "Troca de pressão [Phu]", "Tempo de injeção [ZSx]"]
    resultados = []

    for i, (cp_linha, cpk_linha, pp_linha, ppk_linha) in enumerate(cp_cpk_pp_ppk_linhas):
        parametro = parametros[i]
        try:
            cp = float(cp_linha.split(": ")[1]) if "N/A" not in cp_linha else None
            cpk = float(cpk_linha.split(": ")[1]) if "N/A" not in cpk_linha else None
            pp = float(pp_linha.split(": ")[1]) if "N/A" not in pp_linha else None
            ppk = float(ppk_linha.split(": ")[1]) if "N/A" not in ppk_linha else None
        except ValueError:
            cp, cpk, pp, ppk = None, None, None, None

        if all(v is not None and v >= 1.33 for v in [cp, cpk, pp, ppk]):
            resultado = f"✅ **{parametro}**: O processo está bem controlado e centrado. (Cp: {cp:.2f}, Cpk: {cpk:.2f}, Pp: {pp:.2f}, Ppk: {ppk:.2f})"
        elif all(v is not None and v >= 1.00 for v in [cp, cpk, pp, ppk]):
            resultado = (
                f"<span style='color:#FFA500;'>⚠️ <strong>{parametro}</strong>: Capacidade aceitável, mas pode não estar perfeitamente centrado. "
                f"(Cp: {cp:.2f}, Cpk: {cpk:.2f}, Pp: {pp:.2f}, Ppk: {ppk:.2f})</span>"
            )
        else:
            resultado = (
                f"<span style='color:#FF6F61;'>❌ <strong>{parametro}</strong>: O processo pode precisar de ajustes. "
                f"(Cp: {cp if cp is not None else 'N/A'}, Cpk: {cpk if cpk is not None else 'N/A'}, "
                f"Pp: {pp if pp is not None else 'N/A'}, Ppk: {ppk if ppk is not None else 'N/A'})</span>"
            )

        resultados.append(resultado)

    return resultados


def analisar_resultados_cm_cmk(cm_cmk_linhas):
    """Analisa os resultados de CM e CMK e retorna a análise com cores no Streamlit."""
    parametros = ["Almofada [CPx]", "Troca de pressão [Phu]", "Tempo de injeção [ZSx]"]
    resultados = []

    for i, (cm_linha, cmk_linha) in enumerate(cm_cmk_linhas):
        parametro = parametros[i]
        try:
            cm = float(cm_linha.split(": ")[1]) if "N/A" not in cm_linha else None
            cmk = float(cmk_linha.split(": ")[1]) if "N/A" not in cmk_linha else None
        except ValueError:
            cm, cmk = None, None

        if cm is not None and cmk is not None and cm >= 1.67 and cmk >= 1.67:
            resultado = f"✅ **{parametro}**: A máquina é capaz de produzir dentro das especificações. (CM: {cm:.2f}, CMK: {cmk:.2f})"
        elif cm is not None and cmk is not None and cm >= 1.00 and cmk >= 1.00:
            resultado = (
                f"<span style='color:#FFA500;'>⚠️ <strong>{parametro}</strong>: Capacidade aceitável, mas pode não estar perfeitamente centrado. "
                f"(CM: {cm:.2f}, CMK: {cmk:.2f})</span>"
            )
        else:
            resultado = (
                f"<span style='color:#FF6F61;'>❌ <strong>{parametro}</strong>: A máquina pode precisar de ajustes. "
                f"(CM: {cm if cm is not None else 'N/A'}, CMK: {cmk if cmk is not None else 'N/A'})</span>"
            )

        resultados.append(resultado)

    return resultados






def calcular_indices_formatados(tabela, colunas_limites, parametros, func_calculo, labels):
    resultados = []

    for parametro in parametros:
        try:
            limites = colunas_limites.get(parametro, {'usl': None, 'lsl': None})
            valores = func_calculo(tabela, parametro, limites['usl'], limites['lsl'])
            linha = [f"{label}: {v:.2f}" if v is not None else f"{label}: N/A"
                     for v, label in zip(valores, labels)]
        except Exception:
            linha = [f"{label}: N/A" for label in labels]

        resultados.append(tuple(linha))

    while len(resultados) < len(parametros):
        resultados.append(tuple(f"{label}: N/A" for label in labels))

    return resultados

def exibir_cp_cpk_pp_ppk(tabela, colunas_limites):
    return calcular_indices_formatados(
        tabela,
        colunas_limites,
        ["Almofada [CPx]", "Troca de pressão [Phu]", "Tempo de injeção [ZSx]"],
        lambda df, col, usl, lsl: calcular_cp_cpk(df, col, usl, lsl) + calcular_pp_ppk(df, col, usl, lsl),
        ["Cp", "Cpk", "Pp", "Ppk"]
    )

def exibir_cm_cmk(tabela, colunas_limites):
    return calcular_indices_formatados(
        tabela,
        colunas_limites,
        ["Almofada [CPx]", "Troca de pressão [Phu]", "Tempo de injeção [ZSx]"],
        calcular_cm_cmk,
        ["CM", "CMK"]
    )

def verificar_conformidade_capacidade(analise_cp_cpk, analise_cm_cmk):
    """Verifica se todos os índices estão dentro dos limites aceitáveis."""
    return not any('color:red' in linha for linha in analise_cp_cpk + analise_cm_cmk)

def plotar_grafico_capacidade(ax, tabela, coluna, usl, lsl, tolerancia_superior=None, tolerancia_inferior=None):
    """Plota histogramas dos parâmetros do processo com linhas de especificação e média."""
    cp, cpk = calcular_cp_cpk(tabela, coluna, usl, lsl)
    pp, ppk = calcular_pp_ppk(tabela, coluna, usl, lsl)
    cm, cmk = calcular_cm_cmk(tabela, coluna, usl, lsl) 
    
    # Plot histogram
    ax.hist(tabela[coluna], bins=20, color='blue', edgecolor='black', alpha=0.7, density=True)
    
    # Plot Gaussian curve with wider line
    mean = tabela[coluna].mean()
    std_dev = tabela[coluna].std()
    x = np.linspace(tabela[coluna].min(), tabela[coluna].max(), 100)
    p = norm.pdf(x, mean, std_dev)
    
    ax.plot(x, p, 'k', linewidth=5, label='Curva de Gauss')  # Ajustado para uma espessura mais equilibrada

    # Plot specification lines and mean
    ax.axvline(usl, color='red', linestyle='--', label=f'Limite Sup LSE ({usl:.2f})')
    ax.axvline(lsl, color='red', linestyle='--', label=f'Limite Inf LIE ({lsl:.2f})')
    ax.axvline(mean, color='green', linestyle='--', label=f'Média (X-barra) ({mean:.2f})')
    
    ax.set_title(f'Histograma de {coluna}', fontsize=10, fontweight='bold')
    ax.set_xlabel(coluna, fontsize=8)
    ax.set_ylabel('Frequência', fontsize=8)  # Ajustado para refletir o uso de density=True
    ax.legend(loc='upper left', bbox_to_anchor=(1, 1), fontsize=8)
    ax.grid(True, linestyle='--', alpha=0.5)
    
    # Add Cp, Cpk, Pp, Ppk, Cm, Cmk values to the plot
    cp_text = f'{cp:.2f}' if cp is not None else 'N/A'
    cpk_text = f'{cpk:.2f}' if cpk is not None else 'N/A'
    pp_text = f'{pp:.2f}' if pp is not None else 'N/A'
    ppk_text = f'{ppk:.2f}' if ppk is not None else 'N/A'
    cm_text = f'{cm:.2f}' if cm is not None else 'N/A'
    cmk_text = f'{cmk:.2f}' if cmk is not None else 'N/A'
    
    ax.text(0.95, 0.95, f'Cp: {cp_text}\nCpk: {cpk_text}\nPp: {pp_text}\nPpk: {ppk_text}\nCm: {cm_text}\nCmk: {cmk_text}', 
            transform=ax.transAxes, fontsize=1, verticalalignment='top', horizontalalignment='right', 
            bbox=dict(facecolor='white', alpha=0.5))
    
    return cp, cpk, pp, ppk, cm, cmk

def plotar_grafico(ax, tabela, coluna, titulo, y_label, rebarbas=None, falha=None, cep_superior=None, cep_inferior=None, tolerancia_superior=None, tolerancia_inferior=None):
    """Plota gráficos de tendência para diferentes parâmetros do processo com personalizações."""
    
    # Verifica se a coluna existe e contém dados válidos
    if coluna in tabela.columns and not tabela[coluna].isnull().all():
        tabela[coluna] = pd.to_numeric(tabela[coluna], errors='coerce')
        
        # Verifica se a coluna de injeção existe antes de ordenar e plotar
        if 'Número de injeção [StZx]' in tabela.columns:
            tabela = tabela.sort_values(by='Número de injeção [StZx]', ascending=False)
            eixo_x = tabela['Número de injeção [StZx]']
        else:
            eixo_x = tabela.index  # Usa o índice como fallback
        
        ax.plot(eixo_x, tabela[coluna], marker='o', linestyle='-', color='blue', label=coluna)
        
        # Limites de falha e rebarba
        if rebarbas is not None:
            ax.axhline(rebarbas, color='red', linestyle='--', label=f'Limite Sup LSE ({rebarbas:.2f})')
            ax.fill_between(eixo_x, rebarbas, tabela[coluna].max(), color='red', alpha=0.1)
        if falha is not None:
            ax.axhline(falha, color='red', linestyle='--', label=f'Limite Inf LIE ({falha:.2f})')
            ax.fill_between(eixo_x, tabela[coluna].min(), falha, color='red', alpha=0.1)
        
        # Limites de CEP
        if cep_superior is not None:
            ax.axhline(cep_superior, color='green', linestyle='--', label=f'CEP Superior ({cep_superior:.2f})')
        if cep_inferior is not None:
            ax.axhline(cep_inferior, color='green', linestyle='--', label=f'CEP Inferior ({cep_inferior:.2f})')
        
        # Média
        media = tabela[coluna].mean()
        ax.axhline(media, color='green', linestyle='--', label=f'Média (X-barra) ({media:.2f})')
        
        # Estética
        ax.set_title(titulo, fontsize=10, fontweight='bold')
        ax.set_xlabel('Número de injeção [StZx]' if 'Número de injeção [StZx]' in tabela.columns else 'Índice', fontsize=8)
        ax.set_ylabel(y_label, fontsize=8)
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1), fontsize=8)
        ax.grid(True, linestyle='--', alpha=0.5)
        
        # Limites do eixo X
        if 'Número de injeção [StZx]' in tabela.columns:
            ax.set_xlim(tabela['Número de injeção [StZx]'].min(), tabela['Número de injeção [StZx]'].max())

def salvar_relatorio_pdf(nome_arquivo, caminho_pasta, analise_completa, conforme=True, maquina=None,
                         imagem_largura=20, imagem_altura=15, pos_logo=(170, 10), pos_produto=(10, 10),
                         parametros=None, conforme_capacidade=True):


    def remover_tags_html(texto):
        return re.sub(r'<[^>]+>', '', texto)

    def quebrar_palavras_longa(texto, limite=80):
        palavras = texto.split()
        resultado = []
        for palavra in palavras:
            if len(palavra) > limite:
                partes = textwrap.wrap(palavra, width=limite, break_long_words=True, break_on_hyphens=True)
                partes = [p + '-' if i < len(partes) - 1 else p for i, p in enumerate(partes)]
                resultado.extend(partes)
            else:
                resultado.append(palavra)
        return ' '.join(resultado)

    def limpar_caracteres_invalidos(texto):
        texto = re.sub(r'[\u200B-\u200D\uFEFF]', '', texto)
        texto = re.sub(r'[^\x20-\x7E\n\r\t]', '', texto)
        return texto

    def renderizar_linha_segura(pdf, texto):
        largura_maxima = pdf.w - 2 * pdf.l_margin
        try:
            if pdf.get_string_width(texto) > largura_maxima:
                linhas = textwrap.wrap(texto, width=100)
                for linha in linhas:
                    if pdf.get_string_width(linha) > largura_maxima:
                        pdf.multi_cell(0, 4, "Texto muito longo para renderizar com segurança.")
                    else:
                        pdf.multi_cell(0, 4, linha)
            else:
                pdf.multi_cell(0, 4, texto)
        except Exception:
            try:
                pdf.multi_cell(0, 4, "Texto inválido.")
            except Exception:
                pass

    data_atual = datetime.now().strftime("%Y-%m-%d")
    nome_arquivo_pdf = os.path.splitext(nome_arquivo)[0] + ".pdf"
    caminho_resultado = os.path.join(caminho_pasta, nome_arquivo_pdf)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 7, 'Relatório Monitoramento Parâmetros de Processo', 0, 1, 'C')

    base_imagem_path = "L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo"
    prefixo_arquivo = os.path.basename(nome_arquivo)[:9]
    caminho_imagem_produto = os.path.join(base_imagem_path, "Party Number", f"{prefixo_arquivo}.png")
    caminho_logo = os.path.join(base_imagem_path, "raypro.png")

    if os.path.exists(caminho_imagem_produto):
        try:
            pdf.image(caminho_imagem_produto, x=pos_produto[0], y=pos_produto[1], w=imagem_largura, h=imagem_altura)
        except:
            pdf.cell(0, 5, "Erro ao carregar imagem do produto.", 0, 1, 'C')

    if os.path.exists(caminho_logo):
        try:
            pdf.image(caminho_logo, x=pos_logo[0], y=pos_logo[1], w=25, h=15)
        except:
            pdf.cell(0, 5, "Erro ao carregar logo.", 0, 1, 'C')

    pdf.set_y(max(pos_logo[1] + 15, pos_produto[1] + imagem_altura) + 5)

    pdf.set_font('Helvetica', '', 8)
    nome_formatado = os.path.basename(nome_arquivo).replace('_', ' ').replace('.xlsx', '')
    pdf.cell(0, 5, f"Produto e Ordem de Produção: {nome_formatado}", 0, 1, 'L')
    maquina_formatada = maquina.replace('maquina_', 'Máquina ').replace('_', ' ') if maquina else 'Não Informada'
    pdf.cell(0, 5, f"Máquina: {maquina_formatada}", 0, 1, 'L')
    pdf.cell(0, 5, f"Data do Relatório: {data_atual}", 0, 1, 'L')

    status_capacidade = "Produto Conforme" if conforme_capacidade else "Produto não Conforme"
    pdf.cell(0, 5, f"Status de Capacidade do Processo: {status_capacidade}", 0, 1, 'L')
    pdf.cell(0, 5,
             "Cp, Cpk, Pp, Ppk, Cm e Cmk estão dentro dos limites especificados." if conforme_capacidade
             else "Cp, Cpk, Pp, Ppk, Cm e Cmk estão fora dos limites especificados.",
             0, 1, 'L')

    pdf.set_font('Helvetica', 'B', 8)
    pdf.cell(0, 5, "Resumo dos Valores de Cp, Cpk, Pp, Ppk, Cm e Cmk:", 0, 1, 'L')
    pdf.set_font('Helvetica', '', 7)

    parametros_esperados = ['Almofada', 'Pressão', 'Tempo de Injeção']
    for parametro in parametros_esperados:
        encontrados = [info for info in analise_completa if parametro.lower() in info.lower()]
        if encontrados:
            for info in encontrados:
                texto_limpo = remover_tags_html(info)
                texto_quebrado = quebrar_palavras_longa(texto_limpo, limite=80)
                texto_seguro = limpar_caracteres_invalidos(texto_quebrado)
                renderizar_linha_segura(pdf, texto_seguro)
        else:
            renderizar_linha_segura(pdf, f"{parametro}: Dados não disponíveis.")

    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 7, "Gráfico de Tendência - Histogramas", 0, 1, 'C')

    if os.path.exists('grafico.png'):
        try:
            img = Image.open('grafico.png').convert('RGB')
            img.save('grafico_rgb.png')
            pdf.image('grafico_rgb.png', x=10, y=None, w=190, h=180)
        except:
            pdf.set_font('Helvetica', '', 8)
            pdf.cell(0, 5, "Erro ao carregar gráfico.", 0, 1, 'C')
    else:
        pdf.set_font('Helvetica', '', 8)
        pdf.cell(0, 5, "Gráfico não disponível.", 0, 1, 'C')

    try:
        pdf.output(caminho_resultado)
        print(f"Relatório salvo em: {caminho_resultado}")
    except Exception as e:
        print(f"Erro ao salvar o relatório: {e}")
       
def plotar_grafico_ciclos_por_turno(turno, total_ciclos_monitorados, ciclos_fora_total, percentual_ciclos_fora, total_maquinas_paradas):
    fig = go.Figure()

    # Adicionar barras para cada métrica
    fig.add_trace(go.Bar(
        x=['Ciclos Monitorados', 'Ciclos Fora do Especificado', '% Ciclos Fora do Especificado', 'Paradas de Máquinas'],
        y=[total_ciclos_monitorados, ciclos_fora_total, percentual_ciclos_fora, total_maquinas_paradas],
        text=[total_ciclos_monitorados, ciclos_fora_total, f'{percentual_ciclos_fora:.2f}%', total_maquinas_paradas],
        textposition='auto',
        marker_color=['blue', 'red', 'orange', 'green']
    ))

    # Atualizar layout do gráfico
    fig.update_layout(
        title=f'Monitoramento do Processo - {turno}',
        template='plotly_white'
    )

    return fig

def plotar_grafico_ciclos(total_ciclos_monitorados, ciclos_fora_total, total_maquinas_paradas):
    fig = go.Figure()

    # Adicionar barras para cada métrica
    fig.add_trace(go.Bar(
        x=['Total de Ciclos Monitorados', 'Total de Ciclos Fora do Especificado', 'Total de Paradas de Máquinas'],
        y=[total_ciclos_monitorados, ciclos_fora_total, total_maquinas_paradas],
        text=[total_ciclos_monitorados, ciclos_fora_total, total_maquinas_paradas],
        textposition='auto',
        marker_color=['blue', 'red', 'green']
    ))

    # Atualizar layout do gráfico
    fig.update_layout(
        title='Total Geral - Monitoramento do Processo',
        template='plotly_white'
    )

    return fig

def monitorar_parametros(tabela: pd.DataFrame, parametros: dict) -> tuple:
    global contador_ciclos
    ciclos_fora = 0
    conforme = True
    logs_almofada, logs_pressao, logs_tempo_injecao = [], [], []

    try:
        tabela = filter_last_count(tabela, 'Data')

        parametros_monitoramento = {
            'Almofada [CPx]': (parametros['VALOR_PADRAO_ALMOFADA'], parametros['TOLERANCIA_ALMOFADA'], False),
            'Troca de pressão [Phu]': (parametros['VALOR_PADRAO_PRESSAO'], parametros['TOLERANCIA_PRESSAO'], True),
            'Tempo de injeção [ZSx]': (parametros['VALOR_PADRAO_TEMPO_INJ'], parametros['TOLERANCIA_TEMPO_INJ'], False),
        }

        logging.info("Monitoramento Parâmetros de Processos")
        for parametro, (valor_padrao, tolerancia, percentual) in parametros_monitoramento.items():
            logging.info(f"Parâmetro: {parametro}, Valor Padrão: {valor_padrao}, Tolerância: {tolerancia}, Percentual: {percentual}")
            for index, row in tabela.iterrows():
                valor = row[parametro]
                data = row['Data']

                
                if pd.notna(valor):  # ✅ Ignora valores NaN
                    valor_formatado = f"{valor:.2f}" if parametro != 'Troca de pressão [Phu]' else f"{valor:.3f}"
                    dentro_intervalo = valor_dentro_intervalo(valor, valor_padrao, tolerancia, percentual)
                    log_entry = [data, parametro, valor_formatado, 'Dentro do intervalo' if dentro_intervalo else 'Fora do intervalo']


                    if not dentro_intervalo:
                        ciclos_fora += 1
                        conforme = False

                    if parametro == 'Almofada [CPx]':
                        logs_almofada.append(log_entry)
                    elif parametro == 'Troca de pressão [Phu]':
                        logs_pressao.append(log_entry)
                    elif parametro == 'Tempo de injeção [ZSx]':
                        logs_tempo_injecao.append(log_entry)

        if ciclos_fora > 0:
            contador_ciclos += 1
            logging.info(f"Contador de ciclos: {contador_ciclos} | Ciclos fora: {ciclos_fora}")

        for parametro, (valor_padrao, tolerancia, percentual) in parametros_monitoramento.items():
            tabela[parametro] = tabela[parametro].apply(lambda val: color_negative_red(val, valor_padrao, tolerancia, percentual))

    except KeyError as e:
        logging.error(f"Parâmetro ausente: {e}")
    except Exception as e:
        logging.error(f"Erro ao monitorar parâmetros: {e}", exc_info=True)

    def highlight_out_of_range(row):
        return ['color: red' if row['Status'] == 'Fora do intervalo' else '' for _ in row]

    # ✅ Elimina duplicatas por Data
    df_logs_almofada = pd.DataFrame(logs_almofada, columns=['Data', 'Parâmetro', 'Valor', 'Status']).drop_duplicates(subset=['Data']).style.apply(highlight_out_of_range, axis=1)
    df_logs_pressao = pd.DataFrame(logs_pressao, columns=['Data', 'Parâmetro', 'Valor', 'Status']).drop_duplicates(subset=['Data']).style.apply(highlight_out_of_range, axis=1)
    df_logs_tempo_injecao = pd.DataFrame(logs_tempo_injecao, columns=['Data', 'Parâmetro', 'Valor', 'Status']).drop_duplicates(subset=['Data']).style.apply(highlight_out_of_range, axis=1)

    return conforme, df_logs_almofada, df_logs_pressao, df_logs_tempo_injecao









def main():
    st.markdown('<h1 style="text-align: center;">Monitoramento Parâmetros de Processo</h1>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        st.image("L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/raypro.png", width=80)
    with col3:
        st.image("logo.png", width=250)
    st.markdown('<div class="content">', unsafe_allow_html=True)
    # Lista de máquinas
    maquinas = [
        "maquina_1_48219", "maquina_2_258445", "maquina_3_50765", "maquina_4_258450", "maquina_5_50729",
        "maquina_6_161625", "maquina_7_246628", "maquina_8_246630", "maquina_9_175170", "maquina_10_154998",
        "maquina_11_180433", "maquina_12_163247", "maquina_13_43290", "maquina_14_169626", "maquina_15_46909",
        "maquina_16_200764", "maquina_18_229200", "maquina_20_187905", "maquina_22_216310", "maquina_23_269904",
        "maquina_24_205606", "maquina_25_258448", "maquina_26_219089", "maquina_27_161268", "maquina_28_235399",
        "maquina_29_158040", "maquina_30_252718", "maquina_31_246629", "maquina_32_258446", "maquina_34_258447"
    ]

    # ✅ Definir caminho_parametros antes de usar
    caminho_parametros = r"L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/parametros/parametros.xlsx"
    # ✅ Gerar nomes com status
    if 'maquinas_exibicao_ajustada' not in st.session_state:
        st.session_state.maquinas_exibicao_ajustada = [obter_nome_maquina_com_status(m, caminho_parametros) for m in maquinas]
        st.session_state.maquinas_dict = dict(zip(st.session_state.maquinas_exibicao_ajustada, maquinas))
    
    if 'maquina_atual' not in st.session_state:
        st.session_state.maquina_atual = 0



    # Interface de seleção no sidebar
    with st.sidebar:
        st.image("L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/injetora.png", use_container_width=True)

        maquina_selecionada = st.radio(
            "👉 SELECIONE O EQUIPAMENTO:",
            st.session_state.maquinas_exibicao_ajustada,
            index=st.session_state.maquina_atual,
            key="maquina_radio"
        )

    # Atualiza o índice atual com base na seleção
    st.session_state.maquina_atual = st.session_state.maquinas_exibicao_ajustada.index(maquina_selecionada)

    # Atualiza o status da máquina selecionada
    maquina_real = st.session_state.maquinas_dict[maquina_selecionada]
    novo_status = obter_nome_maquina_com_status(maquina_real, caminho_parametros)

    # Atualiza o nome exibido com novo status
    st.session_state.maquinas_exibicao_ajustada[st.session_state.maquina_atual] = novo_status

    # Atualiza o dicionário de mapeamento
    st.session_state.maquinas_dict = dict(zip(st.session_state.maquinas_exibicao_ajustada, maquinas))

    # Define a variável da máquina selecionada para uso posterior
    maquina = maquina_real











        
    # Placeholder para atualizar a interface
    placeholder = st.empty()
    with placeholder.container():
        # Carregar e processar a tabela
        caminho_pasta = f"L:/Groups\Processos/MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/{maquina}"
        caminho_parametros = r"L:/Groups\Processos/MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/parametros/parametros.xlsx"
        destinatario_email = "leandro.orestes@araymond.com"
        if not os.path.exists(caminho_pasta):
            st.error(f"O caminho especificado não existe: {caminho_pasta}")
            return
        caminho_arquivo = carregar_arquivo(caminho_pasta)
        if caminho_arquivo:
            logging.info(f"Arquivo mais recente carregado com sucesso: {os.path.basename(caminho_arquivo)}")
            logging.info(f"Máquina selecionada: {maquina}")
            tabela = pd.read_excel(caminho_arquivo, engine='openpyxl')
            tabela_processada = processar_tabela(tabela)
            if tabela_processada is not None:
                tabela_ajustada = ajustar_segundos_ciclo(tabela_processada)
                if tabela_ajustada is not None:
                    tabela_sem_duplicatas = remover_duplicatas(tabela_ajustada)
                    if tabela_sem_duplicatas is not None:
                        colunas_para_remover = ['Tempo de dosagem [ZDx]', 'Número de peças boas injetadas [FTZX]', 'Troca de posição [C3U]']
                        tabela_final = remover_colunas(tabela_sem_duplicatas, colunas_para_remover)
                        if tabela_final is not None:
                            tabela_final = tabela_final.reset_index(drop=True)
                            tabela_final = filter_last_count(tabela_final, 'Data')
                            # Calcular a media do tempo de ciclo
                            media_tempo_ciclo = calcular_media_tempo_ciclo(tabela_final)
                            # Verificar paradas de máquina
                            paradas_por_turno = verificar_paradas_maquina(tabela_final, media_tempo_ciclo)
                            for turno, paradas in paradas_por_turno.items():
                                for parada in paradas:
                                    logging.info(f"Parada de máquina detectada em {parada['Data']} com duração de {parada['Duração (minutos)']:.2f} minutos")
                            # Obter o código do produto a partir do nome do arquivo
                            nome_arquivo = os.path.basename(caminho_arquivo)
                            codigo_produto = nome_arquivo.split('_')[0].strip().zfill(9)  # Extrair o código do produto antes do primeiro underscore e garantir 9 dígitos
                            logging.info(f"Código do produto extraído: {codigo_produto}")
                            # Calcular as medias dos parâmetros
                            medias = {
                                'VALOR_PADRAO_TEMPO_INJ': tabela_final['Tempo de injeção [ZSx]'].mean(),
                                'VALOR_PADRAO_ALMOFADA': tabela_final['Almofada [CPx]'].mean(),
                                'VALOR_PADRAO_PRESSAO': tabela_final['Troca de pressão [Phu]'].mean()
                            }

                            # Atualizar a planilha de parâmetros com as médias calculadas
                            atualizar_parametros(caminho_parametros, maquina, codigo_produto, medias)
                            # Carregar os parâmetros do processo
                            parametros = carregar_parametros(caminho_parametros, codigo_produto, maquina)
                            if parametros is None:
                                st.error("Parâmetros de Processos não foram definidos conforme Ficha Técnica. Abortando a execução.")
                                return
                            # Atualizar as constantes com os parâmetros carregados
                            CONSTANTES.update({
                                'VALOR_PADRAO_PRESSAO': parametros['VALOR_PADRAO_PRESSAO'],
                                'VALOR_PRESSAO_REBARBAS': parametros['VALOR_PADRAO_PRESSAO'] + parametros['VALOR_PADRAO_PRESSAO'] * 0.30,  
                                'VALOR_PRESSAO_FALHA': parametros['VALOR_PADRAO_PRESSAO'] - parametros['VALOR_PADRAO_PRESSAO'] * 0.30,  
                                'VALOR_PADRAO_TEMPO_INJ': parametros['VALOR_PADRAO_TEMPO_INJ'],
                                'VALOR_TEMPO_INJ_REBARBAS': parametros['VALOR_PADRAO_TEMPO_INJ'] + parametros['VALOR_PADRAO_TEMPO_INJ'] * 0.15,  
                                'VALOR_TEMPO_INJ_FALHA': parametros['VALOR_PADRAO_TEMPO_INJ'] - parametros['VALOR_PADRAO_TEMPO_INJ'] * 0.15,  
                                'VALOR_PADRAO_ALMOFADA': parametros['VALOR_PADRAO_ALMOFADA'],
                                'VALOR_ALMOFADA_REBARBAS': parametros['VALOR_PADRAO_ALMOFADA'] + parametros['VALOR_PADRAO_ALMOFADA'] * 0.30,  
                                'VALOR_ALMOFADA_FALHA': parametros['VALOR_PADRAO_ALMOFADA'] - parametros['VALOR_PADRAO_ALMOFADA'] * 0.30,  
                                'TOLERANCIA_ALMOFADA': 2.00,  # Valor fixo de 2,00 na planilha
                                'TOLERANCIA_PRESSAO': 0.15,    # Valor fixo de 0,10 na planilha
                                'TOLERANCIA_TEMPO_INJ': 0.15,    # Valor fixo de 0,10 na planilha
                                'LIMITE_ULTIMAS_LINHAS_FORA': parametros['LIMITE_ULTIMAS_LINHAS_FORA']
                            })

                            logging.info("Monitoramento Parâmetros de Processos")
                            conforme, df_logs_almofada, df_logs_pressao, df_logs_tempo_injecao = monitorar_parametros(tabela_final, parametros)

                            # Verificar paradas de máquina
                            paradas_por_turno = verificar_paradas_maquina(tabela_final, media_tempo_ciclo)
                            for turno, paradas in paradas_por_turno.items():
                                for parada in paradas:
                                    logging.info(f"Parada de máquina detectada em {parada['Data']} com duração de {parada['Duração (minutos)']:.2f} minutos")

                            # Formatar as colunas 'Almofada [CPx]' e 'Troca de pressão [Phu]'
                            tabela_final = formatar_coluna_almofada(tabela_final)
                            tabela_final = formatar_coluna_troca_pressao(tabela_final)

                            tab1, tab2, tab3, tab4 = st.tabs(["DADOS RAYPRO", "RESUMO ESTATÍSTICO", "RESUMO PARÂMETROS INJEÇÃO", "HISTÓRICO MÁQUINA"])

                            with tab1:
                               # Adicionar a imagem do produto no canto superior esquerdo e o logo no canto superior direito
                               col1, col2, col3 = st.columns([1, 1, 1])

                               # Caminhos das imagens
                               caminho_imagem_fluxo = "L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/fluxo.png"

                               with tab1:
                                if os.path.exists(caminho_imagem_fluxo):
                                    st.image(caminho_imagem_fluxo, use_container_width=True)
                                else:
                                    st.write("Imagem do fluxo não Cadastrada")

                               # Exibir logs capturados (se houver)
                               st.text(log_capture_string.getvalue())

                               # Exibir a tabela completa sem índice nem contagem adicional
                               st.markdown("### 🗂️🔍 Tabela de Dados Processados RAYPRO")
                               tabela_final = tabela_final.reset_index(drop=True)
                               st.dataframe(tabela_final, use_container_width=True)

                            with tab2:
                               col1, col2, col3 = st.columns([1, 1, 1])
                               prefixo_arquivo = os.path.basename(caminho_arquivo)[:9]
                               caminho_imagem_produto = f"L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/Party Number/{prefixo_arquivo}.png"
                               with col1:
                                   if os.path.exists(caminho_imagem_produto):
                                       st.image(caminho_imagem_produto, width=100)
                                   else:
                                       st.write("Imagem do Produto não Cadastrada")
                               nome_arquivo_formatado = os.path.basename(caminho_arquivo).replace('_', ' ').replace('.xlsx', '')
                               st.text(f"Produto e Ordem de Produção: {nome_arquivo_formatado}")
                               maquina_formatada = f"Máquina {maquina.split('_')[1]} {maquina.split('_')[2]}"
                               st.text(f"Máquina Selecionada: {maquina_formatada}")
                               data_arquivo = time.strftime('%Y-%m-%d', time.gmtime(os.path.getmtime(caminho_arquivo)))
                               st.text(f"Data: {data_arquivo}")
                               st.text(log_capture_string.getvalue())

                               

                               st.markdown("### 🔍 Análise com Verificação de Normalidade")
                               alpha = 0.05  # Valor mínimo aceitável para p-value
                               parametros_indices = {
                                   "Tempo de injeção [ZSx]": ("VALOR_TEMPO_INJ_REBARBAS", "VALOR_TEMPO_INJ_FALHA"),
                                   "Almofada [CPx]": ("VALOR_ALMOFADA_REBARBAS", "VALOR_ALMOFADA_FALHA"),
                                   "Troca de pressão [Phu]": ("VALOR_PRESSAO_REBARBAS", "VALOR_PRESSAO_FALHA")
                               }
                               for parametro, (usl_key, lsl_key) in parametros_indices.items():
                                   usl = parametros.get(usl_key)
                                   lsl = parametros.get(lsl_key)
                                   resultado = calcular_indices_com_normalidade(tabela_final, parametro, usl, lsl)
                                   st.markdown(f"#### 📌 Parâmetro: `{parametro}`")
                                   if resultado is None:
                                       st.error("❌ Dados insuficientes ou limites inválidos para análise estatística.")
                                       continue
                                   tipo = resultado["Distribuição"]
                                   p_value = resultado["p-value"]
                                   st.write(f"**Distribuição dos dados:** `{tipo}`")
                                   if p_value is not None:
                                       st.write(f"**p-value:** `{p_value:.5f}`")
                                       if p_value < alpha:
                                           st.warning(f"⚠️ O p-value está abaixo do limite de significância ({alpha}). Os dados **não são considerados normais**, foram utilizados os índices Pp e Ppk.")
                                       else:
                                           st.success(f"✅ O p-value está acima de {alpha}. Os dados **são considerados normais**, foram utilizados os índices Cp e Cpk.")
                              
                                   

                                   


                               st.markdown("### 📊 Análise dos Resultados de Cp, Cpk, Pp, Ppk")
                               conforme, df_logs_almofada, df_logs_pressao, df_logs_tempo_injecao = monitorar_parametros(tabela, parametros)
                               cp_cpk_linhas = exibir_cp_cpk_pp_ppk(tabela_final, {
                                   'Almofada [CPx]': {'usl': parametros['VALOR_ALMOFADA_REBARBAS'], 'lsl': parametros['VALOR_ALMOFADA_FALHA']},
                                   'Troca de pressão [Phu]': {'usl': parametros['VALOR_PRESSAO_REBARBAS'], 'lsl': parametros['VALOR_PRESSAO_FALHA']},
                                   'Tempo de injeção [ZSx]': {'usl': parametros['VALOR_TEMPO_INJ_REBARBAS'], 'lsl': parametros['VALOR_TEMPO_INJ_FALHA']}
                               })
                               while len(cp_cpk_linhas) < 3:
                                   cp_cpk_linhas.append(("Cp: N/A", "Cpk: N/A", "Pp: N/A", "Ppk: N/A"))
                               analise_cp_cpk = analisar_resultados_cp_cpk_pp_ppk(cp_cpk_linhas)
                               for linha in analise_cp_cpk:
                                   st.markdown(linha, unsafe_allow_html=True) 
                                    
                               st.markdown("### 📈 Análise dos Resultados de Cm e Cmk")
                               conforme, df_logs_almofada, df_logs_pressao, df_logs_tempo_injecao = monitorar_parametros(tabela, parametros)
                               cm_cmk_linhas = exibir_cm_cmk(tabela_final, {
                                   'Almofada [CPx]': {'usl': parametros['VALOR_ALMOFADA_REBARBAS'], 'lsl': parametros['VALOR_ALMOFADA_FALHA']},
                                   'Troca de pressão [Phu]': {'usl': parametros['VALOR_PRESSAO_REBARBAS'], 'lsl': parametros['VALOR_PRESSAO_FALHA']},
                                   'Tempo de injeção [ZSx]': {'usl': parametros['VALOR_TEMPO_INJ_REBARBAS'], 'lsl': parametros['VALOR_TEMPO_INJ_FALHA']}
                               })
                               while len(cm_cmk_linhas) < 3:
                                   cm_cmk_linhas.append(("CM: N/A", "CMK: N/A"))
                               analise_cm_cmk = analisar_resultados_cm_cmk(cm_cmk_linhas)
                               for linha in analise_cm_cmk:
                                   st.markdown(linha, unsafe_allow_html=True)

                               # Verificação final de conformidade
                               st.markdown("### Verificação de Conformidade Final")
                               conforme_capacidade = verificar_conformidade_capacidade(analise_cp_cpk, analise_cm_cmk)
                               st.write(f"Status de Conformidade: {'✅ Produto Conforme' if conforme_capacidade else '❌ Produto não Conforme'}")
                               

                               


                               # --- Gráficos de Tendência e Capacidade ---
                               fig, axs = plt.subplots(3, 2, figsize=(20, 15))
                               fig.subplots_adjust(hspace=0.3, wspace=0.7)

                               # Tendência
                               plotar_grafico(axs[0, 0], tabela_final, 'Tempo de injeção [ZSx]', 'Gráfico de Tendência - Tempo de injeção [ZSx]', 'Tempo de injeção [ZSx]',
                               parametros['VALOR_TEMPO_INJ_REBARBAS'], parametros['VALOR_TEMPO_INJ_FALHA'],
                               cep_superior=parametros.get('CEP_SUPERIOR_TEMPO_INJ'), cep_inferior=parametros.get('CEP_INFERIOR_TEMPO_INJ'),
                               tolerancia_superior=parametros['VALOR_PADRAO_TEMPO_INJ'] + parametros['TOLERANCIA_TEMPO_INJ'],
                               tolerancia_inferior=parametros['VALOR_PADRAO_TEMPO_INJ'] - parametros['TOLERANCIA_TEMPO_INJ'])

                               plotar_grafico(axs[1, 0], tabela_final, 'Almofada [CPx]', 'Gráfico de Tendência - Almofada [CPx]', 'Almofada [CPx]',
                               parametros['VALOR_ALMOFADA_REBARBAS'], parametros['VALOR_ALMOFADA_FALHA'],
                               cep_superior=parametros.get('CEP_SUPERIOR_ALMOFADA'), cep_inferior=parametros.get('CEP_INFERIOR_ALMOFADA'),
                               tolerancia_superior=parametros['VALOR_PADRAO_ALMOFADA'] + parametros['TOLERANCIA_ALMOFADA'],
                               tolerancia_inferior=parametros['VALOR_PADRAO_ALMOFADA'] - parametros['TOLERANCIA_ALMOFADA'])

                               plotar_grafico(axs[2, 0], tabela_final, 'Troca de pressão [Phu]', 'Gráfico de Tendência - Troca de Pressão [Phu]', 'Troca de Pressão [Phu]',
                               parametros['VALOR_PRESSAO_REBARBAS'], parametros['VALOR_PRESSAO_FALHA'],
                               cep_superior=parametros.get('CEP_SUPERIOR_PRESSAO'), cep_inferior=parametros.get('CEP_INFERIOR_PRESSAO'),
                               tolerancia_superior=parametros['VALOR_PADRAO_PRESSAO'] * (1 + parametros['TOLERANCIA_PRESSAO']),
                               tolerancia_inferior=parametros['VALOR_PADRAO_PRESSAO'] * (1 - parametros['TOLERANCIA_PRESSAO']))

                               # Capacidade
                               plotar_grafico_capacidade(axs[0, 1], tabela_final, 'Tempo de injeção [ZSx]',
                               parametros['VALOR_TEMPO_INJ_REBARBAS'], parametros['VALOR_TEMPO_INJ_FALHA'],
                               tolerancia_superior=parametros['VALOR_PADRAO_TEMPO_INJ'] + parametros['TOLERANCIA_TEMPO_INJ'],
                               tolerancia_inferior=parametros['VALOR_PADRAO_TEMPO_INJ'] - parametros['TOLERANCIA_TEMPO_INJ'])

                               plotar_grafico_capacidade(axs[1, 1], tabela_final, 'Almofada [CPx]',
                               parametros['VALOR_ALMOFADA_REBARBAS'], parametros['VALOR_ALMOFADA_FALHA'],
                               tolerancia_superior=parametros['VALOR_PADRAO_ALMOFADA'] + parametros['TOLERANCIA_ALMOFADA'],
                               tolerancia_inferior=parametros['VALOR_PADRAO_ALMOFADA'] - parametros['TOLERANCIA_ALMOFADA'])

                               plotar_grafico_capacidade(axs[2, 1], tabela_final, 'Troca de pressão [Phu]',
                               parametros['VALOR_PRESSAO_REBARBAS'], parametros['VALOR_PRESSAO_FALHA'],
                               tolerancia_superior=parametros['VALOR_PADRAO_PRESSAO'] * (1 + parametros['TOLERANCIA_PRESSAO']),
                               tolerancia_inferior=parametros['VALOR_PADRAO_PRESSAO'] * (1 - parametros['TOLERANCIA_PRESSAO']))

                               # Ajustes visuais
                               for ax in axs.flat:
                                for text in ax.texts:
                                   text.set_fontsize(8)
                                   text.set_bbox(dict(facecolor='white', alpha=0.5, edgecolor='none', boxstyle='round,pad=0.5'))
                                   text.set_position((0.8, 0.8))

                               st.pyplot(fig, use_container_width=True)
                               plt.savefig('grafico.png')

                               # --- Geração de PDF ---
                               def remover_tags_html(texto):
                                import re
                                return re.sub(r'<[^>]+>', '', texto)

                               analise_completa_limpa = [remover_tags_html(linha) for linha in (analise_cp_cpk + analise_cm_cmk)]

                               caminho_resultados = os.path.join(
                               r"L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/Resultados",
                               maquina,
                               "Produto Conforme" if conforme else "Produto não Conforme"
                           )
                               os.makedirs(caminho_resultados, exist_ok=True)

                               salvar_relatorio_pdf(
                               nome_arquivo=os.path.basename(caminho_arquivo),
                               caminho_pasta=caminho_resultados,
                               analise_completa=analise_completa_limpa,
                               conforme=conforme,
                               maquina=maquina,
                               parametros=parametros,
                               conforme_capacidade=conforme_capacidade,
                           )

                               # --- Botão de Download ---
                               caminho_pdf = os.path.join(
                               caminho_resultados,
                               os.path.splitext(os.path.basename(caminho_arquivo))[0] + ".pdf"
                           )
                               with open(caminho_pdf, "rb") as file:
                                   st.download_button("📄 Baixar Relatório PDF", file, file_name=os.path.basename(caminho_pdf))



                                   

                            with tab3:
                                col1, col2, col3 = st.columns([1, 1, 1])
                        
                            
                                def exibir_imagem_produto(prefixo_arquivo, coluna):
                                    caminho = f"L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/Party Number/{prefixo_arquivo}.png"
                                    with coluna:
                                        if os.path.exists(caminho):
                                            st.image(caminho, width=100)
                                        else:
                                            st.write("Imagem do Produto não Cadastrada")

                                exibir_imagem_produto(prefixo_arquivo, col1)

                                nome_arquivo_formatado = os.path.basename(caminho_arquivo).replace('_', ' ').replace('.xlsx', '')
                                st.text(f"Produto e Ordem de Produção: {nome_arquivo_formatado}")

                                maquina_formatada = f"Máquina {maquina.split('_')[1]} {maquina.split('_')[2]}"
                                st.text(f"Máquina Selecionada: {maquina_formatada}")

                                def converter_data(df_estilizado):
                                 """
                                 Converte a coluna 'Data' de um DataFrame:
                                 - Detecta e converte datas em formato serial do Excel
                                 - Garante o tipo datetime
                                 - Trunca microsegundos para segundos
                                 """
                                 try:
                                     df = df_estilizado.data if hasattr(df_estilizado, 'data') else df_estilizado.copy()

                                     if 'Data' in df.columns:
                                         if pd.api.types.is_numeric_dtype(df['Data']):
                                             df['Data'] = pd.to_timedelta(df['Data'], unit='d') + pd.Timestamp('1899-12-30')
                                         df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
                                         df = df.dropna(subset=['Data'])

                                         # Truncar microsegundos para segundos
                                         df['Data'] = df['Data'].dt.floor('s')

                                     return df

                                 except Exception as e:
                                     st.warning(f"Erro ao converter datas: {e}")
                                     return df_estilizado


                                df_logs_almofada = converter_data(df_logs_almofada)
                                df_logs_pressao = converter_data(df_logs_pressao)
                                df_logs_tempo_injecao = converter_data(df_logs_tempo_injecao)

                                # Extrair e formatar a data do arquivo
                                data_arquivo = time.strftime('%Y-%m-%d', time.gmtime(os.path.getmtime(caminho_arquivo)))
                                st.text(f"Data: {data_arquivo}")

                                log_texto = log_capture_string.getvalue()
                                st.text(log_texto[-3000:] if len(log_texto) > 3000 else log_texto)

                                total_ciclos_monitorados_por_turno, ciclos_fora_total, percentual_ciclos_fora_por_turno, ciclos_fora_por_turno, parametros_fora, observacoes_por_turno = verificar_ciclos_por_turno(tabela_final, parametros, media_tempo_ciclo)
                                paradas_por_turno = verificar_paradas_maquina(tabela_final, media_tempo_ciclo)

                                total_maquinas_paradas = sum(len(paradas_por_turno[turno]) for turno in paradas_por_turno)
                                fig_total = plotar_grafico_ciclos(sum(total_ciclos_monitorados_por_turno.values()), ciclos_fora_total, total_maquinas_paradas)

                                col1, col2 = st.columns(2)
                                with col1:
                                    st.plotly_chart(fig_total, use_container_width=True)

                                def plotar_grafico_turno_interativo():
                                    fig = go.Figure()
                                    turnos = ["1° Turno", "2° Turno", "3° Turno"]
                                    metricas = ['Ciclos Monitorados', 'Ciclos Fora do Especificado', '% Ciclos Fora do Especificado', 'Paradas de Máquinas']
                                    cores = ['blue', 'red', 'orange', 'green']

                                    for turno in turnos:
                                        valores = [
                                            total_ciclos_monitorados_por_turno[turno],
                                            ciclos_fora_por_turno[turno],
                                            percentual_ciclos_fora_por_turno[turno],
                                            len(paradas_por_turno[turno])
                                        ]
                                        fig.add_trace(go.Bar(
                                            x=metricas,
                                            y=valores,
                                            name=turno,
                                            marker_color=cores,
                                            visible=(turno == "1° Turno"),
                                            text=[f"{v:.2f}" if isinstance(v, float) else str(v) for v in valores],
                                            textposition='auto'
                                        ))

                                    fig.update_layout(
                                        updatemenus=[{
                                            "buttons": [
                                                {"label": turno,
                                                "method": "update",
                                                "args": [{"visible": [i == j for j in range(3)]},
                                                        {"title": f"Monitoramento do Processo - {turno}"}]}
                                                for i, turno in enumerate(turnos)
                                            ],
                                            "direction": "down",
                                            "showactive": True,
                                            "x": 0.5,
                                            "xanchor": "center",
                                            "y": 1.2,
                                            "yanchor": "top"
                                        }],
                                        title="Monitoramento do Processo - 1° Turno",
                                        template="plotly_white"
                                    )
                                    return fig
                                
                                with col2:
                                    st.plotly_chart(plotar_grafico_turno_interativo(), use_container_width=True)

                                for turno in ["1° Turno", "2° Turno", "3° Turno"]:
                                    st.markdown(f"#### {turno}")
                                    paradas_turno = paradas_por_turno[turno]
    
                                    if paradas_turno:
                                        with st.expander(f"Paradas de Máquina - {turno}"):
                                            for parada in paradas_turno:
                                                st.text(
                                                    f"Parada de Máquina em {parada['Data']} "
                                                    f"com duração de {parada['Duração (minutos)']:.2f} minutos "
                                                    f"({parada['Duração formatada']})"
                                                )
                                    else:
                                        st.text("Nenhuma parada de máquina detectada.")


                                def highlight_out_of_range(row):
                                    return ['color: red' if row['Status'] == 'Fora do intervalo' else '' for _ in row]

                                def preparar_logs(logs, intervalo_minimo_segundos=8):
                                 """
                                 Prepara os logs de parâmetros:
                                 - Converte datas do formato serial Excel para datetime
                                 - Trunca microsegundos para segundos
                                 - Remove valores nulos e zeros
                                 - Elimina duplicatas por Data e Parâmetro
                                 - Remove registros com intervalo menor que o mínimo definido
                                 - Retorna DataFrame limpo sem índice padrão visível
                                 """
                                 df = pd.DataFrame(logs, columns=['Data', 'Parâmetro', 'Valor', 'Status'])

                                 # Converter datas do formato serial Excel
                                 if pd.api.types.is_numeric_dtype(df['Data']):
                                     df['Data'] = pd.to_timedelta(df['Data'], unit='d') + pd.Timestamp('1899-12-30')
                                 df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
                                 df = df.dropna(subset=['Data'])

                                 # Truncar microsegundos para segundos
                                 df['Data'] = df['Data'].dt.floor('s')

                                 # Remover valores nulos e zeros
                                 df = df[df['Valor'].notna()]
                                 df = df[df['Valor'] != 0]

                                 # Eliminar duplicatas por Data e Parâmetro
                                 df = df.drop_duplicates(subset=['Data', 'Parâmetro'], keep='first')

                                 # Ordenar e calcular intervalo entre registros
                                 df = df.sort_values(by='Data')
                                 df['Intervalo'] = df['Data'].diff().dt.total_seconds()

                                 # Remover registros com intervalo menor que o mínimo
                                 df = df[(df['Intervalo'].isna()) | (df['Intervalo'] >= intervalo_minimo_segundos)]

                                 # Excluir coluna Intervalo
                                 df = df.drop(columns=['Intervalo'], errors='ignore')

                                 # Resetar índice para evitar exibição da contagem padrão
                                 df = df.reset_index(drop=True)

                                 # Aplicar destaque visual e ocultar índice
                                 return df.style.apply(highlight_out_of_range, axis=1).hide(axis='index')



                                df_logs_almofada_styled = preparar_logs(df_logs_almofada)
                                df_logs_pressao_styled = preparar_logs(df_logs_pressao)
                                df_logs_tempo_injecao_styled = preparar_logs(df_logs_tempo_injecao)





                                def plotar_grafico_parametro_plotly(tabela, coluna, titulo, y_label, parametros_monitoramento, cor='blue', coluna_ordenacao='Data'):
                                 tabela = tabela.sort_values(by=coluna_ordenacao, ascending=False)
                                 turnos = ["1° Turno", "2° Turno", "3° Turno"]
                                 fig = go.Figure()
                                 traces_por_turno = 4  # Dados + média + tolerância sup + inf

                                 for i, turno in enumerate(turnos):
                                     tabela_turno = tabela[tabela['Turno'] == turno]

                                     # Calcular limites de tolerância
                                     if coluna in parametros_monitoramento:
                                         valor_padrao, tolerancia, percentual = parametros_monitoramento[coluna]
                                         if percentual:
                                             superior = valor_padrao * (1 + tolerancia)
                                             inferior = valor_padrao * (1 - tolerancia)
                                         else:
                                             superior = valor_padrao + tolerancia
                                             inferior = valor_padrao - tolerancia
                                     else:
                                         superior = tabela_turno[coluna].max()
                                         inferior = tabela_turno[coluna].min()

                                     y_min = min(tabela_turno[coluna].min(), inferior)
                                     y_max = max(tabela_turno[coluna].max(), superior)
                                     margem = (y_max - y_min) * 0.1 if y_max != y_min else 1
                                     y_range = [y_min - margem, y_max + margem]

                                     # Trace dos dados
                                     fig.add_trace(go.Scatter(
                                         x=tabela_turno[coluna_ordenacao],
                                         y=tabela_turno[coluna],
                                         mode='lines+markers',
                                         marker=dict(color=cor),
                                         name=f"{coluna}",  # Apenas o nome do parâmetro
                                         visible=(i == 0)
                                     ))

                                     # Linha da média
                                     media_turno = tabela_turno[coluna].mean()
                                     fig.add_trace(go.Scatter(
                                         x=tabela_turno[coluna_ordenacao],
                                         y=[media_turno] * len(tabela_turno),
                                         mode='lines',
                                         line=dict(color='gray', dash='dash'),
                                         name=f"Média - {turno} ({media_turno:.2f})",
                                         visible=(i == 0)
                                     ))

                                     # Tolerância superior
                                     fig.add_trace(go.Scatter(
                                         x=tabela_turno[coluna_ordenacao],
                                         y=[superior] * len(tabela_turno),
                                         mode='lines',
                                         line=dict(color='orange', dash='dot'),
                                         name=f"Tolerância Superior - {turno} ({superior:.2f})",
                                         visible=(i == 0)
                                     ))

                                     # Tolerância inferior
                                     fig.add_trace(go.Scatter(
                                         x=tabela_turno[coluna_ordenacao],
                                         y=[inferior] * len(tabela_turno),
                                         mode='lines',
                                         line=dict(color='orange', dash='dot'),
                                         name=f"Tolerância Inferior - {turno} ({inferior:.2f})",
                                         visible=(i == 0)
                                     ))

                                     # Anotações para pontos fora da tolerância
                                     fora_tolerancia = tabela_turno[(tabela_turno[coluna] > superior) | (tabela_turno[coluna] < inferior)]
                                     for _, row in fora_tolerancia.iterrows():
                                         fig.add_annotation(
                                         x=row[coluna_ordenacao],
                                         y=row[coluna],
                                         text=f"{row[coluna]:.2f}",
                                         showarrow=True,
                                         arrowhead=2,
                                         ax=0,
                                         ay=-20,
                                         font=dict(color="red"),
                                     )

                                 # Menu interativo para alternar entre turnos
                                 fig.update_layout(
                                     updatemenus=[{
                                         "buttons": [
                                             {
                                                 "label": "Todos os Turnos",
                                                 "method": "update",
                                                 "args": [
                                                     {"visible": [True] * (len(turnos) * traces_por_turno)},
                                                     {"title": f"{titulo} - Todos os Turnos"}
                                                 ]
                                             }
                                         ] + [
                                             {
                                                 "label": turno,
                                                 "method": "update",
                                                 "args": [
                                                     {"visible": [j // traces_por_turno == i for j in range(len(turnos) * traces_por_turno)]},
                                                     {"title": f"{titulo} - {turno}"}
                                                 ]
                                             }
                                             for i, turno in enumerate(turnos)
                                         ],
                                         "direction": "down",
                                         "showactive": True,
                                         "x": 0.5,
                                         "xanchor": "center",
                                         "y": 1.2,
                                         "yanchor": "top"
                                     }],
                                     title=f"{titulo} - 1° Turno",
                                     xaxis_title=coluna_ordenacao,
                                     yaxis_title=y_label,
                                     yaxis=dict(range=y_range),
                                     template='plotly_white',
                                     hovermode='closest'
                                 )

                                 return fig
                                
                                # Dicionário de parâmetros com tolerância 
                                parametros_monitoramento = {
                                    'Almofada [CPx]': (parametros['VALOR_PADRAO_ALMOFADA'], parametros['TOLERANCIA_ALMOFADA'], False),
                                    'Troca de pressão [Phu]': (parametros['VALOR_PADRAO_PRESSAO'], parametros['TOLERANCIA_PRESSAO'], True),
                                    'Tempo de injeção [ZSx]': (parametros['VALOR_PADRAO_TEMPO_INJ'], parametros['TOLERANCIA_TEMPO_INJ'], False),
                                }

                                st.markdown("### ⚙️ Almofada")
                                st.dataframe(df_logs_almofada_styled, use_container_width=True)
                                fig_almofada = plotar_grafico_parametro_plotly(
                                    tabela_final,
                                    'Almofada [CPx]',
                                    'Tendência - Almofada [CPx]',
                                    'Almofada [CPx]',
                                    parametros_monitoramento,
                                    cor='blue'
                                )
                                st.plotly_chart(fig_almofada, use_container_width=True)
                                st.markdown("#### 🧪 Pressão de Injeção")
                                st.dataframe(df_logs_pressao_styled, use_container_width=True)
                                fig_pressao = plotar_grafico_parametro_plotly(
                                    tabela_final,
                                    'Troca de pressão [Phu]',
                                    'Tendência - Troca de pressão [Phu]',
                                    'Troca de pressão [Phu]',
                                    parametros_monitoramento,
                                    cor='green'
                                )
                                st.plotly_chart(fig_pressao, use_container_width=True)
                                st.markdown("#### ⏳ Tempo de Injeção")
                                st.dataframe(df_logs_tempo_injecao_styled, use_container_width=True)
                                fig_tempo = plotar_grafico_parametro_plotly(
                                    tabela_final,
                                    'Tempo de injeção [ZSx]',
                                    'Tendência - Tempo de injeção [ZSx]',
                                    'Tempo de injeção [ZSx]',
                                    parametros_monitoramento,
                                    cor='purple'
                                )
                                st.plotly_chart(fig_tempo, use_container_width=True)



                            with tab4:
                                
                                col1, col2, col3 = st.columns([1, 1, 1])

                                prefixo_arquivo = os.path.basename(caminho_arquivo)[:9]
                                caminho_imagem_produto = f"L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/Party Number/{prefixo_arquivo}.png"

                                if os.path.exists(caminho_imagem_produto):
                                    with col1:
                                        st.image(caminho_imagem_produto, width=100)
                                else:
                                    with col1:
                                        st.write("Imagem do Produto não Cadastrada")

                                nome_arquivo_formatado = os.path.basename(caminho_arquivo).replace('_', ' ').replace('.xlsx', '')
                                
                                st.text(f"Produto e Ordem de Produção: {nome_arquivo_formatado}")

                                maquina_formatada = f"Máquina {maquina.split('_')[1]} {maquina.split('_')[2]}"
                                st.text(f"Máquina Selecionada: {maquina_formatada}")

                                # Extrair e formatar a data do arquivo
                                data_arquivo = time.strftime('%Y-%m-%d', time.gmtime(os.path.getmtime(caminho_arquivo)))
                                st.text(f"Data: {data_arquivo}")

                                st.text(log_capture_string.getvalue())

                                st.title("📊 Estatística Produto x Máquina")

                                # Carregar o DataFrame
                                try: 
                                    df = pd.read_excel("L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/resultados_estatisticos.xlsx", engine="openpyxl")
                                except Exception as e:
                                    st.error(f"Erro ao carregar o histórico estatístico: {e}")
                                    st.stop()

                                # Função de conformidade completa
                                def calcular_conformidade_completa(df, maquina, part_number):
                                    part_number = str(part_number).zfill(9)
                                    df_filtrado = df[
                                        (df["Maquina"] == maquina) &
                                        (df["Part Number"].astype(str).str.zfill(9) == part_number)
                                    ]
                                    total = len(df_filtrado)
                                    if total == 0:
                                        return "Sem dados para cálculo de conformidade."

                                    parametros = ["Tempo_Inj", "Almofada", "Pressao"]
                                    indices = ["Cp", "Cpk", "Pp", "Ppk", "Cm", "Cmk"]
                                    colunas = [f"{i}_{p}" for i in indices for p in parametros]
                                    colunas_existentes = [c for c in colunas if c in df_filtrado.columns]
                                    limites = {"Cp": 1.00, "Cpk": 1.00, "Pp": 1.00, "Ppk": 1.00,  "Cm": 1.00, "Cmk": 1.00}

                                    def linha_conforme(row):
                                        for col in colunas_existentes:
                                            indice = col.split("_")[0]
                                            if pd.isna(row[col]) or row[col] < limites[indice]:
                                                return False
                                        return True

                                    qtd_ok = df_filtrado.apply(linha_conforme, axis=1).sum()
                                    qtd_nok = total - qtd_ok
                                    porcentagem_ok = (qtd_ok / total) * 100
                                    porcentagem_nok = (qtd_nok / total) * 100

                                    return (
                                        f"Total de Processos Monitorados: {total}\n"
                                        f"Quantidade Dentro do Especificado: {qtd_ok}\n"
                                        f"Quantidade Fora do Especificado: {qtd_nok}\n"
                                        f"Porcentagem de Conformidade: {porcentagem_ok:.2f}%\n"
                                        f"Porcentagem de Não Conformidade: {porcentagem_nok:.2f}%"
                                    )

                                # Exibir resultado
                                st.markdown("**Resumo Estatístico Completo:**")
                                st.text(calcular_conformidade_completa(df, maquina, prefixo_arquivo))

                                try:
                                 df = pd.read_excel("resultados_estatisticos.xlsx", engine="openpyxl")
                                except Exception as e:
                                    st.error(f"Erro ao carregar o histórico estatístico: {e}")
                                    st.stop()

                                # Filtro por máquina e produto atual
                                df_maquina = df[
                                    (df["Maquina"] == maquina) &  
                                    (df["Part Number"].astype(str).str.zfill(9) == prefixo_arquivo)
                                ].copy()

                                if df_maquina.empty:
                                    st.warning("Não há dados estatísticos disponíveis para esta máquina e produto.")
                                    st.stop()

                                # ✅ Remover '.0' da coluna 'Ordem de Produção'
                                if "Ordem de Produção" in df_maquina.columns:
                                    df_maquina["Ordem de Produção"] = df_maquina["Ordem de Produção"].astype(str).str.replace(r'\.0$', '', regex=True)

                                df_maquina["Arquivo"] = df_maquina["Part Number"].astype(str) + "\n" + df_maquina["Ordem de Produção"].astype(str)
                                df_maquina = df_maquina.sort_values(by="Data", ascending=False)

                                indice_selecionado = st.radio("Selecione o índice para visualização:",
                                                               options=["Cp e Cpk", "Pp e Ppk", "Cm e Cmk"],
                                                               horizontal=True)

                                colunas_necessarias = {
                                    "Cp e Cpk": ["Cp_Tempo_Inj", "Cpk_Tempo_Inj", "Cp_Almofada", "Cpk_Almofada", "Cp_Pressao", "Cpk_Pressao"],
                                    "Pp e Ppk": ["Pp_Tempo_Inj", "Ppk_Tempo_Inj", "Pp_Almofada", "Ppk_Almofada", "Pp_Pressao", "Ppk_Pressao"],
                                    "Cm e Cmk": ["Cm_Tempo_Inj", "Cmk_Tempo_Inj", "Cm_Almofada", "Cmk_Almofada", "Cm_Pressao", "Cmk_Pressao"]
                                }

                                faltantes = [col for col in colunas_necessarias[indice_selecionado] if col not in df_maquina.columns]
                                if faltantes:
                                    st.warning(f"Colunas ausentes para visualização: {', '.join(faltantes)}")
                                    st.stop()

                                def create_bar_chart(df, y_columns, title, limite_conformidade, limite_aceitavel=None):
                                 df['Media_Indice'] = df[y_columns].mean(axis=1)
                                 df = df.sort_values(by='Media_Indice', ascending=False)

                                 fig = px.bar(
                                     df,
                                     x="Arquivo",
                                     y=y_columns,
                                     title=title,
                                     labels={"value": "", "variable": "", "Arquivo": ""}
                                 )

                                 # Linha vermelha do Limite de Conformidade
                                 fig.add_hline(
                                     y=limite_conformidade,
                                     line_dash="dash",
                                     line_color="red",
                                     annotation_text=f"Limite de Conformidade {y_columns[0].split('_')[0]}",
                                     annotation_position="top left"
                                 )

                                 # Linha amarela indicando Limite mínimo aceitável
                                 if limite_aceitavel is not None:
                                     fig.add_hline(
                                         y=limite_aceitavel,
                                         line_dash="dot",
                                         line_color="orange",
                                         annotation_text="Limite mínimo aceitável",
                                         annotation_position="bottom left"
                                     )

                                 fig.update_layout(
                                     xaxis=dict(
                                         tickmode='array',
                                         tickvals=df["Arquivo"],
                                         ticktext=df["Arquivo"],
                                         tickangle=-45,
                                         title=""
                                     ),
                                     yaxis_title="",
                                     legend_title_text=""
                                 )
                                 return fig

                                if indice_selecionado == "Cp e Cpk":
                                    fig1 = create_bar_chart(df_maquina, ["Cp_Tempo_Inj", "Cpk_Tempo_Inj"], "Cp/Cpk - Tempo de Injeção", 1.33, 1.00)
                                    fig2 = create_bar_chart(df_maquina, ["Cp_Almofada", "Cpk_Almofada"], "Cp/Cpk - Almofada", 1.33, 1.00)
                                    fig3 = create_bar_chart(df_maquina, ["Cp_Pressao", "Cpk_Pressao"], "Cp/Cpk - Pressão", 1.33, 1.00)
                                elif indice_selecionado == "Pp e Ppk":
                                    fig1 = create_bar_chart(df_maquina, ["Pp_Tempo_Inj", "Ppk_Tempo_Inj"], "Pp/Ppk - Tempo de Injeção", 1.33, 1.00)
                                    fig2 = create_bar_chart(df_maquina, ["Pp_Almofada", "Ppk_Almofada"], "Pp/Ppk - Almofada", 1.33, 1.00)
                                    fig3 = create_bar_chart(df_maquina, ["Pp_Pressao", "Ppk_Pressao"], "Pp/Ppk - Pressão", 1.33, 1.00)
                                else:
                                    fig1 = create_bar_chart(df_maquina, ["Cm_Tempo_Inj", "Cmk_Tempo_Inj"], "Cm/Cmk - Tempo de Injeção", 1.67, 1.00)
                                    fig2 = create_bar_chart(df_maquina, ["Cm_Almofada", "Cmk_Almofada"], "Cm/Cmk - Almofada", 1.67, 1.00)
                                    fig3 = create_bar_chart(df_maquina, ["Cm_Pressao", "Cmk_Pressao"], "Cm/Cmk - Pressão", 1.67, 1.00)

                                # Após exibir os gráficos com st.plotly_chart(...)
                                col1, col2, col3 = st.columns([1, 1, 1])  

                                with col1:
                                    st.plotly_chart(fig1, use_container_width=True)
                                with col2:
                                    st.plotly_chart(fig2, use_container_width=True)
                                with col3:
                                    st.plotly_chart(fig3, use_container_width=True)

                                # Legenda explicativa
                                st.markdown("""
                                <div style='padding-top: 10px; font-size: 14px;'>
                                🔴 <span style='color:red;'>**Linha vermelha tracejada**</span>: Limite de Conformidade (≥ 1.33 para Cp/Cpk e Pp/Ppk, ≥ 1.67 para Cm/Cmk)<br>
                                🟠 <span style='color:orange;'>**Linha amarela pontilhada**</span>: Limite mínimo aceitável (≥ 1.00 para Cp/Cpk e Pp/Ppk e Cm/Cmk)<br>
                                📊 Os valores abaixo (< 1.00 para Cp/Cpk e Pp/Ppk e Cm/Cmk) indicam necessidade de ajustes no Processo.
                                </div>
                                """, unsafe_allow_html=True)






                                # Função para Seleção na aba lateral
                                if modo == "Histórico Estatístico MU.PLÁSTICOS":
                                 historico_estatistico_interface()

                                caminho_parametros = r"L:/Groups\Processos/MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/parametros/parametros.xlsx"
                                # Carregar e atualizar histórico estatístico de todas as máquinas
                                if 'historico_dados_geral' not in st.session_state:
                                    with st.spinner('Carregando dados de todas as máquinas...'):
                                        st.session_state['historico_dados_geral'] = carregar_dados_todas_maquinas(maquinas, caminho_parametros)

                                        # Atualizar o histórico com os dados gerais (apenas uma vez)
                                        for linha in st.session_state['historico_dados_geral']:
                                         atualizar_resultados_estatisticos(CAMINHO_HISTORICO, linha)

                                        # ✅ Atualizar os nomes exibidos com os novos status após o processamento
                                        st.session_state.maquinas_exibicao_ajustada = [
                                            obter_nome_maquina_com_status(maquina, caminho_parametros) for maquina in maquinas
                                        ]
                                        st.session_state.maquinas_dict = dict(zip(st.session_state.maquinas_exibicao_ajustada, maquinas))
                                        st.success("Status das máquinas atualizado com base nos dados mais recentes.")

                                # Salvar os resultados estatísticos no histórico
                                nome_arquivo = os.path.basename(caminho_arquivo) 

                                # Extrair Part Number e Ordem de Produção do nome do arquivo
                                if "_" in nome_arquivo and nome_arquivo.endswith(".xlsx"):
                                    part_number, ordem_producao = nome_arquivo.replace(".xlsx", "").split("_", 1)
                                else:
                                    part_number, ordem_producao = "", ""

                                # Calcular os totais antes de construir nova_linha
                                total_ciclos_monitorados = sum(total_ciclos_monitorados_por_turno.values())
                                total_maquinas_paradas = sum(len(paradas_por_turno[turno]) for turno in paradas_por_turno)

                                # Construção do dicionário nova_linha
                                nova_linha = {
                                    "Maquina": maquina,
                                    "Arquivo": nome_arquivo,
                                    "Part Number": part_number.lstrip("0"),
                                    "Ordem de Produção": ordem_producao,
                                    "Cp_Tempo_Inj": cp_cpk_linhas[2][0].split(": ")[1] if "N/A" not in cp_cpk_linhas[2][0] else None,
                                    "Cpk_Tempo_Inj": cp_cpk_linhas[2][1].split(": ")[1] if "N/A" not in cp_cpk_linhas[2][1] else None,
                                    "Pp_Tempo_Inj": cp_cpk_linhas[2][2].split(": ")[1] if "N/A" not in cp_cpk_linhas[2][2] else None,
                                    "Ppk_Tempo_Inj": cp_cpk_linhas[2][3].split(": ")[1] if "N/A" not in cp_cpk_linhas[2][3] else None,
                                    "Cm_Tempo_Inj": cm_cmk_linhas[2][0].split(": ")[1] if "N/A" not in cm_cmk_linhas[2][0] else None,
                                    "Cmk_Tempo_Inj": cm_cmk_linhas[2][1].split(": ")[1] if "N/A" not in cm_cmk_linhas[2][1] else None,
                                    "Cp_Almofada": cp_cpk_linhas[0][0].split(": ")[1] if "N/A" not in cp_cpk_linhas[0][0] else None,
                                    "Cpk_Almofada": cp_cpk_linhas[0][1].split(": ")[1] if "N/A" not in cp_cpk_linhas[0][1] else None,
                                    "Pp_Almofada": cp_cpk_linhas[0][2].split(": ")[1] if "N/A" not in cp_cpk_linhas[0][2] else None,
                                    "Ppk_Almofada": cp_cpk_linhas[0][3].split(": ")[1] if "N/A" not in cp_cpk_linhas[0][3] else None,
                                    "Cm_Almofada": cm_cmk_linhas[0][0].split(": ")[1] if "N/A" not in cm_cmk_linhas[0][0] else None,
                                    "Cmk_Almofada": cm_cmk_linhas[0][1].split(": ")[1] if "N/A" not in cm_cmk_linhas[0][1] else None,
                                    "Cp_Pressao": cp_cpk_linhas[1][0].split(": ")[1] if "N/A" not in cp_cpk_linhas[1][0] else None,
                                    "Cpk_Pressao": cp_cpk_linhas[1][1].split(": ")[1] if "N/A" not in cp_cpk_linhas[1][1] else None,
                                    "Pp_Pressao": cp_cpk_linhas[1][2].split(": ")[1] if "N/A" not in cp_cpk_linhas[1][2] else None,
                                    "Ppk_Pressao": cp_cpk_linhas[1][3].split(": ")[1] if "N/A" not in cp_cpk_linhas[1][3] else None,
                                    "Cm_Pressao": cm_cmk_linhas[1][0].split(": ")[1] if "N/A" not in cm_cmk_linhas[1][0] else None,
                                    "Cmk_Pressao": cm_cmk_linhas[1][1].split(": ")[1] if "N/A" not in cm_cmk_linhas[1][1] else None,
                                    "Total_Ciclos_Monitorados": total_ciclos_monitorados,
                                    "Total_Ciclos_Fora": ciclos_fora_total,
                                    "Total_Paradas_Maquina": total_maquinas_paradas
                                }

                                # Verificação dos valores estatísticos antes de atualizar o histórico
                                colunas_estatisticas = [
                                    "Cp_Tempo_Inj", "Cpk_Tempo_Inj", "Pp_Tempo_Inj", "Ppk_Tempo_Inj", "Cm_Tempo_Inj", "Cmk_Tempo_Inj",
                                    "Cp_Almofada", "Cpk_Almofada", "Pp_Almofada", "Ppk_Almofada", "Cm_Almofada", "Cmk_Almofada",
                                    "Cp_Pressao", "Cpk_Pressao", "Pp_Pressao", "Ppk_Pressao", "Cm_Pressao", "Cmk_Pressao"
                                ]

                                if all(nova_linha.get(col) in [None, ""] for col in colunas_estatisticas):
                                    st.info(f"Linha ignorada por ausência de valores estatísticos: {nome_arquivo}")
                                else:
                                    atualizar_resultados_estatisticos(CAMINHO_HISTORICO, nova_linha)

                                # Lista de máquinas
                                maquinas = [
                                    "maquina_1_48219", "maquina_2_258445", "maquina_3_50765", "maquina_4_258450",
                                    "maquina_5_50729", "maquina_6_161625", "maquina_7_246628", "maquina_8_246630",
                                    "maquina_9_175170", "maquina_10_154998", "maquina_11_180433", "maquina_12_163247",
                                    "maquina_13_43290", "maquina_14_169626", "maquina_15_46909", "maquina_16_200764",
                                    "maquina_18_229200", "maquina_20_187905", "maquina_22_216310", "maquina_23_269904", "maquina_24_205606",
                                    "maquina_25_258448", "maquina_26_219089", "maquina_27_161268", "maquina_28_235399",
                                    "maquina_29_158040", "maquina_30_252718", "maquina_31_246629", "maquina_32_258446",
                                    "maquina_34_258447"
                                ]

CAMINHO_HISTORICO = "L:/Groups/Processos/MU PLÁSTICO/AUTOMAÇÃO DE TAREFAS/Monitoramento Parâmetros de Processo/resultados_estatisticos.xlsx"

def carregar_historico_estatistico(caminho_excel):
    if os.path.exists(caminho_excel):
        try:
            return pd.read_excel(caminho_excel, engine="openpyxl", dtype={"Part Number": str})
        except Exception as e:
            st.error(f"Erro ao carregar histórico: {e}")
    return pd.DataFrame()

def destacar_valores(df):
    df = df.copy()

    # Seleciona apenas colunas numéricas
    colunas_numericas = df.select_dtypes(include='number').columns

    # Aplica estilo com duas casas decimais apenas nas colunas numéricas
    styled = df.style.format({col: "{:.2f}" for col in colunas_numericas})

    # Aqui você pode adicionar outras regras de estilo, como destaque de valores fora do padrão
    # Exemplo:
    # styled = styled.applymap(lambda v: 'background-color: red' if v < 1 else '', subset=colunas_numericas)

    return styled

def exportar_dataframe_para_excel(df):
    """
    Exporta um DataFrame para um arquivo Excel em memória (BytesIO).
    Retorna os bytes prontos para download.
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return output

def formatar_nome_maquina(maquina):
    partes = maquina.split('_')
    if len(partes) > 2:
        return f"Máquina {partes[1]} {partes[2]}"
    return maquina

def formatar_nome_arquivo(nome):
    nome_formatado = nome.replace('_', ' ').replace('.xlsx', '')
    codigo_produto = nome_formatado.split(' ')[0]
    return f"{codigo_produto}\n{nome_formatado[len(codigo_produto)+1:]}"

def gerar_grafico(tipo, x, y, texto, titulo, ticktext=None):
    if tipo == "Barras":
        fig = go.Figure(data=[
            go.Bar(
                x=x,
                y=y,
                marker_color='#0066CC',  # vermelho suave
                text=texto,
                textposition='auto'
            )
        ])
    elif tipo == "Linha":
        fig = go.Figure(data=[
            go.Scatter(
                x=x,
                y=y,
                mode='lines+markers',
                text=texto
            )
        ])
    else:
        fig = go.Figure()

    fig.update_layout(title=titulo)

    if ticktext:
        fig.update_layout(xaxis=dict(
            tickmode='array',
            tickvals=list(range(len(ticktext))),
            ticktext=ticktext,
            tickangle=60
        ))

    return fig

# 🧩 Constantes de nomes de colunas
COL_CICLOS_MONITORADOS = "Total_Ciclos_Monitorados"
COL_CICLOS_FORA = "Total_Ciclos_Fora"
COL_PARADAS_MAQUINA = "Total_Paradas_Maquina"

# 🔧 Funções principais
def total_ciclos_monitorados(df, maquina, part_number):
    df_filtrado, erro = df(df, maquina, part_number)
    if erro:
        return erro
    if COL_CICLOS_MONITORADOS not in df_filtrado.columns:
        return f"Coluna '{COL_CICLOS_MONITORADOS}' não encontrada."
    total = pd.to_numeric(df_filtrado[COL_CICLOS_MONITORADOS], errors="coerce").sum()
    return f"Total de Registros: {len(df_filtrado)}\nTotal de Ciclos Monitorados: {int(total)}"

def total_ciclos_fora(df, maquina, part_number):
    df_filtrado, erro = (df, maquina, part_number)
    if erro:
        return erro
    if COL_CICLOS_FORA not in df_filtrado.columns:
        return f"Coluna '{COL_CICLOS_FORA}' não encontrada."
    total = pd.to_numeric(df_filtrado[COL_CICLOS_FORA], errors="coerce").sum()
    return f"Total de Ciclos Fora do Especificado: {int(total)}"

def total_paradas_maquina(df, maquina, part_number):
    df_filtrado, erro = df(df, maquina, part_number)
    if erro:
        return erro
    if COL_PARADAS_MAQUINA not in df_filtrado.columns:
        return f"Coluna '{COL_PARADAS_MAQUINA}' não encontrada."
    total = pd.to_numeric(df_filtrado[COL_PARADAS_MAQUINA], errors="coerce").sum()
    return f"Total de Paradas de Máquina: {int(total)}"

# 📊 Média por Part Number
def calcular_media_por_part_number(df):
    colunas = [COL_CICLOS_MONITORADOS, COL_CICLOS_FORA, COL_PARADAS_MAQUINA]
    for coluna in colunas:
        if coluna in df.columns:
            df[coluna] = pd.to_numeric(df[coluna], errors="coerce")
        else:
            df[coluna] = 0
    return df.groupby("Part Number")[colunas].mean().round(2)

def calcular_conformidade_completa(
    df,
    maquina=None,
    part_number=None,
    ordem_producao=None,
    data_inicio=None,
    data_fim=None,
    indices_selecionados=None,
    parametros_selecionados=None,
    **kwargs
):
    df_filtrado = df.copy()

    # 🔍 Filtros básicos
    if maquina:
        df_filtrado = df_filtrado[df_filtrado["Maquina"] == maquina]

    if part_number:
        part_number = str(part_number).zfill(9)
        df_filtrado = df_filtrado[df_filtrado["Part Number"].astype(str).str.zfill(9) == part_number]

    if ordem_producao:
        df_filtrado = df_filtrado[df_filtrado["Ordem de Produção"] == ordem_producao]
        
    # 🗓️ Filtro por data
    if "Data" in df_filtrado.columns:
        df_filtrado["Data"] = pd.to_datetime(df_filtrado["Data"], errors="coerce")
        if data_inicio:
            df_filtrado = df_filtrado[df_filtrado["Data"] >= pd.to_datetime(data_inicio)]
        if data_fim:
            df_filtrado = df_filtrado[df_filtrado["Data"] <= pd.to_datetime(data_fim)]

    # 📉 Verificação de conformidade
    total = len(df_filtrado)
    if total == 0:
        return "Sem dados para cálculo de conformidade."

    indices = indices_selecionados or ["Cp", "Cpk", "Pp", "Ppk", "Cm", "Cmk"]
    parametros = parametros_selecionados or ["Tempo_Inj", "Almofada", "Pressao"]
    colunas = [f"{i}_{p}" for i in indices for p in parametros]
    colunas_existentes = [c for c in colunas if c in df_filtrado.columns]
    limites = {indice: 1.00 for indice in indices}

    def linha_conforme(row):
        for col in colunas_existentes:
            indice = col.split("_")[0]
            try:
                valor = float(row[col])
                if pd.isna(valor) or valor < limites[indice]:
                    return False
            except (ValueError, TypeError):
                return False
        return True

    qtd_ok = df_filtrado.apply(linha_conforme, axis=1).sum()
    qtd_nok = total - qtd_ok
    porcentagem_ok = (qtd_ok / total) * 100
    porcentagem_nok = (qtd_nok / total) * 100
    media_ciclos_monitorados = pd.to_numeric(df_filtrado[COL_CICLOS_MONITORADOS], errors="coerce").mean()
    media_ciclos_fora = pd.to_numeric(df_filtrado[COL_CICLOS_FORA], errors="coerce").mean()
    media_paradas_maquina = pd.to_numeric(df_filtrado[COL_PARADAS_MAQUINA], errors="coerce").mean()

    return (
        f"Total de Processos Monitorados: {total}\n"
        f"Quantidade Dentro do Especificado: {qtd_ok}\n"
        f"Quantidade Fora do Especificado: {qtd_nok}\n"
        f"Porcentagem de Conformidade: {porcentagem_ok:.2f}%\n"
        f"Porcentagem de Não Conformidade: {porcentagem_nok:.2f}%\n"
        f"Média de Ciclos Monitorados por Processo: {media_ciclos_monitorados:.0f}\n"
        f"Média de Ciclos Fora por Processo: {media_ciclos_fora:.0f}\n"
        f"Média de Paradas de Máquina por Processo: {media_paradas_maquina:.0f}"
    )

# ✅ Função que separa os dados em conformes e não conformes
def calcular_df_conformidade(df, indices_selecionados=None, parametros_selecionados=None):
    indices = indices_selecionados or ["Cp", "Cpk", "Pp", "Ppk", "Cm", "Cmk"]
    parametros = parametros_selecionados or ["Tempo_Inj", "Almofada", "Pressao"]
    colunas = [f"{i}_{p}" for i in indices for p in parametros]
    colunas_existentes = [c for c in colunas if c in df.columns]

    condicoes = pd.Series(False, index=df.index)
    for col in colunas_existentes:
        condicoes |= pd.to_numeric(df[col], errors='coerce') < 1.00

    df_fora = df[condicoes]
    df_dentro = df[~condicoes]
    return df_fora, df_dentro

def historico_estatistico_interface():
    st.title("📈 Histórico Estatístico MU.PLÁSTICOS")
    df_hist = carregar_historico_estatistico(CAMINHO_HISTORICO)

    if df_hist.empty:
        st.warning("Nenhum dado estatístico encontrado.")
        return

    colunas_esperadas = [
        'Maquina', 'Part Number', 'Ordem de Produção', 'Data',
        'Cp_Tempo_Inj', 'Cpk_Tempo_Inj', 'Pp_Tempo_Inj', 'Ppk_Tempo_Inj', 'Cm_Tempo_Inj', 'Cmk_Tempo_Inj',
        'Cp_Almofada', 'Cpk_Almofada', 'Pp_Almofada', 'Ppk_Almofada', 'Cm_Almofada', 'Cmk_Almofada',
        'Cp_Pressao', 'Cpk_Pressao', 'Pp_Pressao', 'Ppk_Pressao', 'Cm_Pressao', 'Cmk_Pressao',
        'Total_Ciclos_Monitorados', 'Total_Ciclos_Fora', 'Total_Paradas_Maquina'
    ]
    colunas_faltantes = [col for col in colunas_esperadas if col not in df_hist.columns]
    if colunas_faltantes:
        st.error(f"Colunas ausentes no histórico: {', '.join(colunas_faltantes)}")
        return

    df_hist['Data'] = pd.to_datetime(df_hist['Data'], errors='coerce')
    df_hist = df_hist.dropna(subset=['Data'])

    col1, col2, col3 = st.columns(3)

    with col1:
        filtro_maquina = st.multiselect("Filtrar por Máquina", sorted(df_hist['Maquina'].dropna().unique()))
        filtro_part = st.multiselect("Filtrar por Part Number", sorted(df_hist['Part Number'].dropna().unique()))

    with col2:
        if 'Ordem de Produção' in df_hist.columns:
           df_hist['Ordem de Produção'] = df_hist['Ordem de Produção'].astype(str).str.replace(r'\.0$', '', regex=True)

        filtro_ordem = st.multiselect("Filtrar por Ordem de Produção", sorted(df_hist['Ordem de Produção'].dropna().unique()))
        data_min = df_hist['Data'].min()
        data_max = df_hist['Data'].max()
        filtro_data = st.date_input("Filtrar por intervalo de datas", value=(data_min, data_max), min_value=data_min, max_value=data_max)

    with col3:
        indice_estatistico = st.multiselect("Filtrar por Índice Estatístico", options=["Cp", "Cpk", "Pp", "Ppk", "Cm", "Cmk"])
        parametro_injecao = st.multiselect("Filtrar por Parâmetro de Injeção", options=["Tempo_Inj", "Almofada", "Pressao"])

    df_filtrado = aplicar_filtros(df_hist.copy(), filtro_maquina, filtro_part, filtro_ordem, filtro_data, indice_estatistico, parametro_injecao)
    df_filtrado['Data'] = pd.to_datetime(df_filtrado['Data'], errors='coerce').dt.date
    df_filtrado = reordenar_colunas(df_filtrado)
    df_exibicao = df_filtrado.drop(columns=['Arquivo'], errors='ignore')

    if 'Ordem de Produção' in df_exibicao.columns:
        df_exibicao['Ordem de Produção'] = df_exibicao['Ordem de Produção'].apply(
            lambda x: str(int(x)) if pd.notna(x) and isinstance(x, (int, float)) else str(x)
        )

    colunas_para_arredondar = df_exibicao.select_dtypes(include='number').columns.difference(['Ordem de Produção'])
    df_exibicao[colunas_para_arredondar] = df_exibicao[colunas_para_arredondar].round(2)
    df_estilizado = destacar_valores(df_exibicao)

    st.subheader("📋 Tabela de Resultados")
    st.dataframe(df_estilizado, use_container_width=True)

    excel_bytes = exportar_dataframe_para_excel(df_filtrado)
    st.download_button(
        "📥 Baixar dados filtrados em Excel",
        data=excel_bytes,
        file_name="historico_estatistico_filtrado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.markdown("**📊 Resumo Estatístico Completo:**")
    data_inicio, data_fim = filtro_data if isinstance(filtro_data, tuple) and all(filtro_data) else (None, None)
    maquina = filtro_maquina[0] if filtro_maquina else None
    part_number = filtro_part[0] if filtro_part else None
    ordem_producao = filtro_ordem[0] if filtro_ordem else None

    st.text(calcular_conformidade_completa(
        df_hist,
        maquina=maquina,
        part_number=part_number,
        ordem_producao=ordem_producao,
        data_inicio=data_inicio,
        data_fim=data_fim,
        indices_selecionados=indice_estatistico,
        parametros_selecionados=parametro_injecao
    ))

    df_fora, df_dentro = calcular_df_conformidade(df_filtrado, indices_selecionados=indice_estatistico, parametros_selecionados=parametro_injecao)
    tipo_grafico = st.selectbox("Tipo de gráfico", ["Barras", "Linha"])

    # ✅ Gráfico por Máquina com filtro aplicado
    df_contagem_fora = df_fora.groupby('Maquina').size().reset_index(name='Fora')
    df_contagem_dentro = df_dentro.groupby('Maquina').size().reset_index(name='Dentro')
    df_contagem = pd.merge(df_contagem_fora, df_contagem_dentro, on='Maquina', how='outer').fillna(0)
    df_contagem['Total'] = df_contagem['Fora'] + df_contagem['Dentro']
    df_contagem['% Fora'] = (df_contagem['Fora'] / df_contagem['Total']) * 100
    df_contagem = df_contagem[df_contagem['Total'] > 0]

    if filtro_maquina:
        df_contagem = df_contagem[df_contagem['Maquina'].isin(filtro_maquina)]

    df_contagem['Maquina'] = df_contagem['Maquina'].apply(formatar_nome_maquina)
    df_contagem = df_contagem[df_contagem['Maquina'].notna() & (df_contagem['Maquina'].astype(str).str.strip() != '')]
    df_contagem = df_contagem.sort_values(by='% Fora', ascending=False)

    fig = exibir_grafico(tipo_grafico, df_contagem['Maquina'], df_contagem['% Fora'],
                         df_contagem['% Fora'].apply(lambda x: f'{x:.2f}%'),
                         '📉 % Produtos Fora do Especificado por Máquina (Cp,Cpk,Pp,Ppk,Cm e Cmk)')
    fig.update_layout(xaxis_tickangle=-45, margin=dict(l=40, r=40, t=60, b=120))
    st.plotly_chart(fig, use_container_width=True)

    # ✅ Gráfico TOP 10 Produtos com filtro aplicado
    if 'Part Number' in df_fora.columns:
        df_fora['Grupo_Arquivo'] = df_fora['Part Number'].astype(str).str.zfill(9)

        if filtro_part:
            df_fora = df_fora[df_fora['Grupo_Arquivo'].isin(filtro_part)]

        top10 = df_fora['Grupo_Arquivo'].value_counts().head(10).reset_index()
        top10.columns = ['Grupo_Arquivo', 'Fora']
        top10['Indice'] = range(len(top10))

        ticktext = [formatar_nome_arquivo(nome) for nome in top10['Grupo_Arquivo']]
        fig2 = exibir_grafico(tipo_grafico, top10['Indice'], top10['Fora'], top10['Fora'],
                              '🔍 TOP 10 Produtos Fora do Especificado (Cp,Cpk,Pp,Ppk,Cm e Cmk)', ticktext=ticktext)
        fig2.update_layout(xaxis_tickangle=-45, margin=dict(l=40, r=40, t=60, b=120))
        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.warning("A coluna 'Part Number' não está presente no DataFrame df_fora.")

    # ✅ Adiciona o gráfico mensal
    exibir_grafico_mensal(df_filtrado, filtro_maquina=filtro_maquina, filtro_part=filtro_part)

def aplicar_filtros(
    df,
    filtro_maquina=None,
    filtro_part=None,
    filtro_ordem=None,
    filtro_data=None,
    indice_estatistico=None,
    parametro_injecao=None
):
    df_filtrado = df.copy()

    # 🧭 Filtros básicos
    if filtro_maquina:
        df_filtrado = df_filtrado[df_filtrado['Maquina'].isin(filtro_maquina)]

    if filtro_ordem:
        df_filtrado = df_filtrado[df_filtrado['Ordem de Produção'].isin(filtro_ordem)]

    if filtro_part:
        df_filtrado = df_filtrado[df_filtrado['Part Number'].isin(filtro_part)]

    # 📅 Filtro por intervalo de datas
    if filtro_data and isinstance(filtro_data, tuple) and len(filtro_data) == 2:
        df_filtrado['Data'] = pd.to_datetime(df_filtrado['Data'], errors='coerce')
        df_filtrado = df_filtrado[
            (df_filtrado['Data'] >= pd.to_datetime(filtro_data[0])) &
            (df_filtrado['Data'] <= pd.to_datetime(filtro_data[1]))
        ]

    # 🧠 Filtro por colunas específicas de índice estatístico e parâmetro de injeção
    if indice_estatistico or parametro_injecao:
        colunas_desejadas = []
        indices = indice_estatistico or ["Cp", "Cpk", "Pp", "Ppk", "Cm", "Cmk"]
        parametros = parametro_injecao or ["Tempo_Inj", "Almofada", "Pressao"]
        for i in indices:
            for p in parametros:
                coluna = f"{i}_{p}"
                if coluna in df_filtrado.columns:
                    colunas_desejadas.append(coluna)

        # Mantém apenas as colunas desejadas + principais
        colunas_principais = ['Maquina', 'Part Number', 'Ordem de Produção', 'Data', 'Arquivo']
        colunas_existentes = [col for col in colunas_principais if col in df_filtrado.columns]
        df_filtrado = df_filtrado[colunas_existentes + colunas_desejadas]

    return df_filtrado

def reordenar_colunas(df):
    colunas = df.columns.tolist()
    if 'Data' in colunas and 'Maquina' in colunas:
        colunas.remove('Data')
        index_maquina = colunas.index('Maquina')
        colunas.insert(index_maquina, 'Data')
        df = df[colunas]
    return df

def calcular_df_conformidade(df, indices_selecionados=None, parametros_selecionados=None):
    indices = indices_selecionados or ["Cp", "Cpk", "Pp", "Ppk", "Cm", "Cmk"]
    parametros = parametros_selecionados or ["Tempo_Inj", "Almofada", "Pressao"]
    colunas = [f"{i}_{p}" for i in indices for p in parametros]
    colunas_existentes = [c for c in colunas if c in df.columns]

    condicoes = pd.Series(False, index=df.index)
    for col in colunas_existentes:
        condicoes |= pd.to_numeric(df[col], errors='coerce') < 1.00

    df_fora = df[condicoes]
    df_dentro = df[~condicoes]
    return df_fora, df_dentro

def exibir_grafico(tipo, x, y, texto, titulo, ticktext=None):
    return gerar_grafico(tipo, x, y, texto, titulo, ticktext=ticktext)

# Tenta definir o locale para português do Brasil
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')  # Windows
    except locale.Error:
        pass  # Se não conseguir, segue com o padrão do sistema

# Tenta definir o locale para português do Brasil
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')  # Windows
    except locale.Error:
        pass  # Se não conseguir, segue com o padrão do sistema

    



def exibir_grafico_mensal(df_filtrado, filtro_maquina=None, filtro_part=None):
    # Aplicar filtros se fornecidos
    if filtro_maquina:
        df_filtrado = df_filtrado[df_filtrado['Maquina'].isin(filtro_maquina)]
    if filtro_part:
        df_filtrado = df_filtrado[df_filtrado['Part Number'].astype(str).str.zfill(9).isin(filtro_part)]

    # Converter datas
    df_filtrado['Data'] = pd.to_datetime(df_filtrado['Data'], errors='coerce')
    df_filtrado = df_filtrado.dropna(subset=['Data'])
    df_filtrado['Mes'] = df_filtrado['Data'].dt.strftime('%B').str.capitalize()
    df_filtrado['MesNum'] = df_filtrado['Data'].dt.month

    # Verificar conformidade
    def determinar_status(row):
        for col in row.index:
            if any(metric in col for metric in ['Cp', 'Cpk', 'Pp', 'Ppk', 'Cm', 'Cmk']):
                try:
                    if float(row[col]) < 1.00:
                        return 'Fora do Especificado'
                except:
                    continue
        return 'Dentro do Especificado'

    df_filtrado['Status'] = df_filtrado.apply(determinar_status, axis=1)

    # Agrupar por mês e status
    df_mensal = df_filtrado.groupby(['MesNum', 'Mes', 'Status']).size().reset_index(name='Quantidade')
    df_mensal = df_mensal.sort_values('MesNum')

    # Criar gráfico
    fig_mensal = px.bar(
        df_mensal,
        x='Mes',
        y='Quantidade',
        color='Status',
        color_discrete_map={
            'Dentro do Especificado': '#87CEFA',
            'Fora do Especificado': '#0066CC'
        },
        barmode='group',
        text='Quantidade',
        title='📆 Quantidade Mensal de Processos Dentro e Fora do Especificado'
    )

    fig_mensal.update_traces(textposition='inside', textfont_color='white')
    fig_mensal.update_layout(
        xaxis_title=None,
        yaxis_title=None,
        xaxis_tickangle=-25,
        margin=dict(l=20, r=20, t=60, b=50)
    )

    st.plotly_chart(fig_mensal, use_container_width=True)





    












    
  
    
   

# Raypro Automate
import pyautogui
import pandas as pd
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import win32com.client as win32
import logging
import schedule
import time
import matplotlib.pyplot as plt
import numpy as np
import streamlit as st
import os
import pyperclip
import subprocess
import threading
import ctypes  
import platform


# Variável global de controle para parar a automação
stop_event = threading.Event()

# Configuração do logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Sistema de segurança
pyautogui.FAILSAFE = True

# Caminho das imagens
caminho_imagens = r"L:\Groups\Processos\Processos\RAYPRO AUTOMATE\IMAGENS"

def localizar_e_clicar(imagem, descricao, duracao=1, confidence=0.85, tentativas=5, tempo_espera=2):
    """Localiza e clica em uma imagem na tela com tentativas múltiplas e ajuste dinâmico de confiança."""
    tentativa_atual = 0
    while tentativa_atual < tentativas:
        try:
            logging.info(f"Tentando localizar a imagem: {imagem} (Tentativa {tentativa_atual + 1} de {tentativas}) com confiança {confidence}")
            local = pyautogui.locateOnScreen(os.path.join(caminho_imagens, imagem), confidence=confidence)
            if local:
                x, y = pyautogui.center(local)
                logging.info(f"Posição do mouse antes de clicar: ({x}, {y})")
                pyautogui.moveTo(x, y, duration=duracao)
                pyautogui.click()
                logging.info(f"{descricao} localizada e clicada.")
                time.sleep(1)
                return True
            else:
                logging.warning(f"Imagem '{imagem}' não encontrada na tentativa {tentativa_atual + 1}.")
                tentativa_atual += 1
                confidence -= 0.05
                time.sleep(tempo_espera)
        except Exception as e:
            logging.error(f"Erro ao tentar localizar a imagem '{imagem}': {e}")
            tentativa_atual += 1
            time.sleep(tempo_espera)
    logging.error(f"Falha ao localizar a imagem '{imagem}' após {tentativas} tentativas.")
    pyautogui.screenshot(f"erro_{imagem}.png")
    return False

def localizar_e_clicar_thread(imagem, descricao, duracao=1, confidence=0.85, tentativas=5, tempo_espera=2):
    """Localiza e clica em uma imagem na tela em uma thread separada."""
    thread = threading.Thread(target=localizar_e_clicar, args=(imagem, descricao, duracao, confidence, tentativas, tempo_espera))
    thread.start()
    return thread

def registrar_posicao_mouse():
    """Registra a posição atual do mouse."""
    x, y = pyautogui.position()
    logging.info(f"Posição atual do mouse: ({x}, {y})")
    return x, y

def localizar_e_clicar_com_registro(imagem, descricao, duracao=1, confidence=0.85, tentativas=5):
    """Localiza e clica em uma imagem na tela, registrando a posição do mouse para evitar erros."""
    if localizar_e_clicar(imagem, descricao, duracao, confidence, tentativas):
        x, y = registrar_posicao_mouse()
        logging.info(f"Imagem '{imagem}' localizada e clicada na posição ({x}, {y}).")
        return True
    else:
        logging.error(f"Falha ao localizar e clicar na imagem '{imagem}'.")
        return False

def abrir_raypro():
    directory = "C:\\Program Files (x86)\\Raynet\\Raypro_Supervision\\RayProSupervision.exe"
    try:
        if platform.system() == "Windows":
            # Cria um novo console para o RayPro, isolando do terminal principal
            subprocess.Popen([directory], creationflags=subprocess.CREATE_NEW_CONSOLE)
        else:
            subprocess.Popen([directory])
        logging.info("RayPro aberto via subprocess.")
        time.sleep(10)
        return True
    except Exception as e:
        logging.error(f"Erro ao abrir RayPro: {e}")
        return False

def abrir_mapa_do_supervisor(x=697, y=393):
    """Abre o Mapa do Supervisor clicando diretamente na posição especificada."""
    logging.info("Abrindo Mapa do Supervisor pela posição do mouse.")
    time.sleep(10) # Aguarda o sistema carregar
    pyautogui.moveTo(x, y, duration=1)
    pyautogui.click()
    logging.info(f"Mapa do Supervisor clicado na posição ({x}, {y}).")

def abrir_brazil_fst_inj(x=1087, y=245):
    logging.info("Abrindo BRAZIL FST INJ pela posição do mouse.")
    time.sleep(10)
    pyautogui.moveTo(x, y, duration=1)
    pyautogui.click()
    logging.info(f"BRAZIL FST INJ clicado na posição ({x}, {y}).")

def localizar_maquina(x, y, nome_maquina):
    logging.info(f"Localizando {nome_maquina}.")
    time.sleep(8)
    pyautogui.moveTo(x, y, duration=1)
    pyautogui.click()

    logging.info(f"{nome_maquina} localizada na posição ({x}, {y})! ⚙ ")


def abrir_poogle():
    logging.info("Abrindo POOGLE.")
    if localizar_e_clicar_com_registro("poogle.png", "POOGLE"):
        logging.info("POOGLE aberto com sucesso.")
        return True
    else:
        logging.error("Falha ao abrir POOGLE. Tentando fechar a janela com Alt + F4.")
        pyautogui.hotkey('alt', 'f4')
        time.sleep(2)
        return False

def localizar_produto(x=425, y=240):
    logging.info("Localizando produto.")
    time.sleep(5)
    pyautogui.moveTo(x, y, duration=1)
    pyautogui.click()
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('ctrl', 'c')
    produto = pyperclip.paste().strip()
    if produto:
        produto = produto[:9].replace(" ", "")
        logging.info(f"Produto copiado: {produto}")
        return produto
    else:
        logging.error("Falha ao localizar o produto.")
        return None

def localizar_ordemprod(x=233, y=240):
    logging.info("Localizando ordem de produção.")
    time.sleep(5)
    pyautogui.moveTo(x, y, duration=1)
    pyautogui.click()
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('ctrl', 'c')
    ordemprod = pyperclip.paste().strip()
    if ordemprod:
        ordemprod = ordemprod[:8]
        logging.info(f"Ordem de Produção copiada: {ordemprod}")
        return ordemprod
    else:
        logging.error("Falha ao localizar a ordem de produção.")
        return None

def localizar_parametro():
    logging.info("Localizando parâmetro.")
    time.sleep(5)
    localizar_e_clicar_com_registro("parameter.png", "Parâmetro")

def exportar_para_excel(caminho_pasta, produto, ordemprod):
    logging.info("Exportando para Excel.")
    time.sleep(15)
    try:
        if not os.path.exists(caminho_pasta):
            os.makedirs(caminho_pasta)
            logging.info(f"Diretório criado: {caminho_pasta}")

        if localizar_e_clicar_com_registro("exportador.png", "Exportador", duracao=15, confidence=0.8):
            nome_arquivo = f"{produto}_{ordemprod}.xlsx".replace("-", "_")
            caminho_completo = os.path.join(caminho_pasta, nome_arquivo)

            for arquivo in os.listdir(caminho_pasta):
                caminho_arquivo = os.path.join(caminho_pasta, arquivo)
                if os.path.isfile(caminho_arquivo):
                    os.remove(caminho_arquivo)
                    logging.info(f"Arquivo antigo removido: {caminho_arquivo}")

            pyperclip.copy(caminho_completo)
            pyautogui.hotkey("ctrl", "v")
            pyautogui.press("enter")
            time.sleep(2)
            pyautogui.press("enter")
            logging.info(f"Arquivo salvo em {caminho_completo} e substituído automaticamente.")
        else:
            logging.error("Falha ao localizar o exportador.")
    except Exception as e:
        logging.error(f"Erro ao exportar para Excel: {e}")

def fechar_exportador(x=614, y=57):
    logging.info("Fechando exportador.")
    time.sleep(5)
    pyautogui.moveTo(x, y, duration=1)
    pyautogui.click()

def processar_maquina(nome_maquina, posicao_maquina, caminho_pasta):
    if stop_event.is_set():
        logging.info("Botão de parada pressionado antes do início. Encerrando RayPro.")
        fechar_raypro()
        return

    logging.info(f"Processando {nome_maquina}.")
    localizar_maquina(*posicao_maquina, nome_maquina)
    if not abrir_poogle():
        logging.error("Abortando processamento da máquina devido à falha ao abrir POOGLE.")
        return
    produto = localizar_produto()
    if not produto:
        logging.error("Falha ao localizar o produto.")
        return
    ordemprod = localizar_ordemprod()
    if not ordemprod:
        logging.error("Falha ao localizar a ordem de produção.")
        return
    localizar_parametro()
    exportar_para_excel(caminho_pasta, produto, ordemprod)
    fechar_exportador()

    if stop_event.is_set():
        logging.info("Botão de parada pressionado. Encerrando RayPro.")
        fechar_raypro()

def fechar_raypro():
    if stop_event.is_set():
        logging.info("Fechando RayPro.")
        time.sleep(5)
        pyautogui.hotkey('alt', 'f4')
        logging.info("RayPro fechado automaticamente.")
    else:
        logging.info("Stop event não acionado. RayPro continuará aberto.")


    # Definir as máquinas e suas posições
maquinas = [
    {'nome': 'maquina_1_48219', 'posicao': (36, 228), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_1_48219'},
    {'nome': 'maquina_2_258445', 'posicao': (354, 218), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_2_258445'},
    {'nome': 'maquina_3_50765', 'posicao': (589, 212), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_3_50765'},
    {'nome': 'maquina_4_258450', 'posicao': (818, 228), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_4_258450'},
    {'nome': 'maquina_5_50729', 'posicao': (1049, 214), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_5_50729'},
    {'nome': 'maquina_6_161625', 'posicao': (1280, 215), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_6_161625'},
    {'nome': 'maquina_7_246628', 'posicao': (126, 334), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_7_246628'},
    {'nome': 'maquina_8_246630', 'posicao': (362, 330), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_8_246630'},
    {'nome': 'maquina_9_175170', 'posicao': (588, 346), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_9_175170'},
    {'nome': 'maquina_10_154998', 'posicao': (816, 340), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_10_154998'},
    {'nome': 'maquina_11_180433', 'posicao': (1062, 342), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_11_180433'},
    {'nome': 'maquina_12_163247', 'posicao': (1277, 342), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_12_163247'},
    {'nome': 'maquina_13_43290', 'posicao': (138, 476), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_13_43290'},
    {'nome': 'maquina_14_169626', 'posicao': (366, 471), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_14_169626'},
    {'nome': 'maquina_15_46909', 'posicao': (581, 459), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_15_46909'},
    {'nome': 'maquina_16_200764', 'posicao': (822, 473), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_16_200764'},
    {'nome': 'maquina_18_229200', 'posicao': (1065, 487), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_18_229200'},
    {'nome': 'maquina_20_187905', 'posicao': (1282, 482), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_20_187905'},
    {'nome': 'maquina_22_216310', 'posicao': (121, 605), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_22_216310'},
    {'nome': 'maquina_23_269904', 'posicao': (352, 611), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_23_269904'},
    {'nome': 'maquina_24_205606', 'posicao': (596, 601), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_24_205606'},
    {'nome': 'maquina_25_258448', 'posicao': (822, 597), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_25_258448'},
    {'nome': 'maquina_26_219089', 'posicao': (1058, 602), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_26_219089'},
    {'nome': 'maquina_27_161268', 'posicao': (1290, 593), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_27_161268'},
    {'nome': 'maquina_28_235399', 'posicao': (108, 753), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_28_235399'},
    {'nome': 'maquina_29_158040', 'posicao': (349, 750), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_29_158040'},
    {'nome': 'maquina_30_252718', 'posicao': (590, 739), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_30_252718'},
    {'nome': 'maquina_31_246629', 'posicao': (804, 732), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_31_246629'},
    {'nome': 'maquina_32_258446', 'posicao': (1064, 746), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_32_258446'},
    {'nome': 'maquina_34_258447', 'posicao': (1293, 748), 'caminho_pasta': r'\\br.ray.group\world_br$\Groups\Processos\MU PLÁSTICO\AUTOMAÇÃO DE TAREFAS\Monitoramento Parâmetros de Processo\maquina_34_258447'},
]



def baixar_arquivos():
    logging.info("Iniciando processo de baixar arquivos.")
    if abrir_raypro():
        abrir_mapa_do_supervisor() 
        abrir_brazil_fst_inj()
        for maquina in maquinas:
            if stop_event.is_set():
                logging.info("Parada solicitada. Interrompendo processamento das máquinas.")
                break
            processar_maquina(maquina["nome"], maquina["posicao"], maquina["caminho_pasta"])
        fechar_raypro()
    else:
        logging.error("Falha ao abrir RayPro.")

    # ✅ Garantia extra de encerramento
    if stop_event.is_set():
        logging.info("Botão de parada pressionado. Encerrando RayPro.")
        fechar_raypro()


def executar_raypro_automate():
    logging.info("Automação Raypro iniciada.")
    try:
        while not stop_event.is_set():
            baixar_arquivos()
            # Verifica novamente após baixar_arquivos
            if stop_event.is_set():
                logging.info("Parada detectada após ciclo. Encerrando automação.")
                break
            time.sleep(1)
    except Exception as e:
        logging.error(f"Erro durante execução do Raypro Automate: {e}")
    finally:
        logging.info("Automação Raypro encerrada com sucesso.")







if __name__ == "__main__":
    with st.sidebar:

        modo = st.radio(
            "👉 SELECIONE A OPERAÇÃO:",
            ["MONITORAMENTO DE PROCESSO", "HISTÓRICO ESTATÍSTICO", "RAYPRO AUTOMATE"]
        )
    if modo == "MONITORAMENTO DE PROCESSO":
        main()
    elif modo == "HISTÓRICO ESTATÍSTICO":
        historico_estatistico_interface()
    elif modo == "RAYPRO AUTOMATE":
        executar_raypro_automate()




















        

























